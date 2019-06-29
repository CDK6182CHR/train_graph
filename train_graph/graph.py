"""
运行图类.
同时管理self._config和self._sysConfig两套设置系统，两套系统完全分离。
"""
from .line import Line
from .ruler import Ruler
from .train import Train
from .circuit import Circuit, CircuitNode
from copy import copy
from Timetable_new.utility import stationEqual
import json, re
from datetime import datetime
from .pyETRCExceptions import *

config_file = 'config.json'

import cgitb

cgitb.enable(format='text')


class Graph:
    """
    运行图类，数据结构：
    Line line;
    List<Train> _trains=[];
    List<Circuit> _circuit;//交路数据
    Dict _config;//系统设置，主要是UI
    str markdown;//附注
    """

    def __init__(self):
        """
        构造空类，不考虑读文件
        """
        self.filename = ""
        self._version = ""
        self.line = Line()
        self._trains = []
        self._circuits = []
        self._config = None
        self._sysConfig = None
        self.readSysConfig()  # 初始化并校验系统默认配置文件
        self.typeList = []  # public data
        self.initGraphConfig()
        self._markdown = ''
        self.fullCheciMap = {}  # 全车次查找表 str|->Train
        self.singleCheciMap = {}  # 单车次查找表str|->list<Train>

    def readSysConfig(self):
        """
        1.4新定义函数。读取并检查系统默认设置。
        """
        try:
            with open(config_file, encoding='utf-8', errors='ignore') as fp:
                self._sysConfig = json.load(fp)
        except:
            print("配置文件config.json加载失败，使用系统默认设置。")
            self._sysConfig = {}
        self.checkSysConfig()

    def checkSysConfig(self):
        """
        1.4版本新增函数。每次在系统设置中增加新的字段时，要修改本函数。
        """
        default_config = {
            "seconds_per_pix": 15.0,
            "seconds_per_pix_y": 8.0,
            "pixes_per_km": 4.0,
            "grid_color": "#00FF00",
            "text_color": "#0000FF",
            "default_keche_width": 1.5,
            "default_huoche_width": 0.75,
            "start_hour": 0,
            "end_hour": 24,
            "minutes_per_vertical_line": 10.0,
            "bold_line_level": 2,
            "show_line_in_station": True,
            "start_label_height": 30,
            "end_label_height": 15,
            "table_row_height": 30,
            "link_line_height":10,
            "show_time_mark": 1,  # 显示详细停点。0-不显示，1-仅显示选中车次，2-显示所有车次
            "max_passed_stations": 3,  # 至多跨越站数。超过这个数将被分成两段运行图。
            "default_colors": {"快速": "#FF0000", "特快": "#0000FF",
                               "直达特快": "#FF00FF", "动车组": "#804000", "动车": "#804000",
                               "高速": "#FF00BE", "城际": "#FF33CC", "default": "#008000"
                               },
            "margins": {
                "left_white": 15,  # 左侧白边，不能有任何占用的区域
                "right_white": 10,
                "left": 325,
                "up": 90,
                "down": 90,
                "right": 170,
                "label_width": 100,
                "mile_label_width": 50,
                "ruler_label_width": 100,
            },
            "type_regex": [
                ('高速', r'G\d+', True),
                ('动车组', r'D\d+', True),
                ('城际', r'C\d+', True),
                ('直达特快', r'Z\d+', True),
                ('特快', r'T\d+', True),
                ('快速', r'K\d+', True),
                ('普快', r'[1-5]\d{3}$', True),  # 非复用
                ('普快', r'[1-5]\d{3}\D', True),  # 复用
                ('普客', r'6\d{3}$', True),
                ('普客', r'6\d{3}\D', True),
                ('普客', r'7[1-5]\d{2}$', True),
                ('普客', r'7[1-5]\d{2}\D', True),
                ('通勤', r'7\d{3}$', True),
                ('通勤', r'7\d{3}\D', True),
                ('通勤', r'8\d{3}$', True),
                ('通勤', r'8\d{3}\D', True),
                ('旅游', r'Y\d+', True),
                ('路用', r'57\d+', True),
                ('特快行包', r'X1\d{2}', True),
                ('动检', r'DJ\d+', True),
                ('客车底', r'0[GDCZTKY]\d+', True),
                ('临客', r'L\d+', True),  # 主要解决类型直接定为“临客”的车次。
                ('客车底', r'0\d{4}', True),
                ('行包', r'X\d{3}\D', False),
                ('行包', r'X\d{3}$', False),
                ('班列', r'X\d{4}', False),
                ('直达', r'1\d{4}', False),
                ('直货', r'2\d{4}', False),
                ('区段', r'3\d{4}', False),
                ('摘挂', r'4[0-4]\d{3}', False),
                ('小运转', r'4[5-9]\d{3}', False),
                ('单机', r'5[0-2]\d{3}', False),
                ('补机', r'5[3-4]\d{3}', False),
                ('试运转', r'55\d{3}', False),
            ]  # 类型对应正则表达式，list<tuple>有序三元列表。数据结构为类型，正则，是否客车。优先级递减。
        }
        default_config.update(self._sysConfig)
        self._sysConfig = default_config

    def saveSysConfig(self):
        """
        1.4版本新增函数。保证进入本函数时系统设置是有效的。
        """
        with open(config_file, 'w', encoding='utf-8', errors='ignore') as fp:
            json.dump(self._sysConfig, fp, ensure_ascii=False)

    def initGraphConfig(self):
        """
        1.4新增函数。在不考虑读文件的前提下，增加单个运行图设置应该比系统运行图设置多的东西。
        precondition: sysConfig已经读取并校验完毕。
        """
        self._config = {
            "ordinate": None,
            "not_show_types": [],
        }
        self._config.update(self._sysConfig)

    def checkGraphConfig(self):
        """
        1.4新增函数，由loadGraph调用（初始化时可以用更加暴力的方法）。
        检查并更新graphConfig。precondition: self._config存在且是dict；sysConfig已经初始化并检查完毕。
        """
        self._config.setdefault("ordinate", None)
        self._config.setdefault("not_show_types", [])
        for key, value in self._sysConfig.items():
            self._config.setdefault(key, value)

    def resetGraphConfigFromConfigWidget(self):
        """
        1.4版本新增函数。提供给configWidget调用的重置接口。
        precondition: self._config和self._sysConfig都是合法的。
        """
        # for key,value in self._sysConfig:
        #     if 'color' not in key:
        #         self._config[key]=value
        self._config.update(self._sysConfig)

    def setFullCheciMap(self):
        """
        设置全车次查找表。
        """
        for train in self.trains():
            self.fullCheciMap[train.fullCheci()] = train

    def setSingleCheciMap(self):
        for train in self.trains():
            for cc in (train.downCheci(), train.upCheci()):
                if cc:
                    self.singleCheciMap.setdefault(cc, []).append(train)

    def addSingleCheciMap(self, train):
        for cc in (train.downCheci(), train.upCheci()):
            if cc:
                self.singleCheciMap.setdefault(cc, []).append(train)

    def delSingleCheciMap(self, train):
        for cc in (train.downCheci(), train.upCheci()):
            if cc:
                lst = self.singleCheciMap.get(cc)
                if len(lst) > 1:
                    lst.remove(train)
                else:
                    del self.singleCheciMap[cc]

    def changeTrainCheci(self, train: Train, full: str, down: str, up: str):
        """
        2019.02.24新增加，更新车次，并更新查找表。
        """
        f, d, u = train.checi
        if full != f and self.fullCheciMap.get(f, None) is not None:
            del self.fullCheciMap[f]
            self.fullCheciMap[full] = train
        reset = False
        if d != down or u != up:
            self.delSingleCheciMap(train)
            reset = True
        train.setCheci(full, down, up)
        if reset:
            self.addSingleCheciMap(train)

    def refreshData(self):
        """
        2019.02.24新增。系统刷新时强制刷新所有映射表数据，保证正常工作。
        """
        self.setFullCheciMap()
        self.setSingleCheciMap()
        self.line.setNameMap()
        self.line.setFieldMap()

    def loadGraph(self, filename: str):
        """
        暂定直接打开json文件读
        """
        self.filename = filename
        fp = open(filename, encoding='utf8', errors='ignore')
        try:
            info = json.load(fp)
        except json.JSONDecodeError:
            fp.close()
            self.loadTrcGraph(filename)
            return

        self.line.loadLine(info["line"])
        self._circuits = []
        for c in info.get('circuits', []):
            self._circuits.append(Circuit(self, origin=c))
        for dict_train in info["trains"]:
            newtrain = Train(self, origin=dict_train)
            self._trains.append(newtrain)

        self._config = info.get("config", {})
        if not isinstance(self._config, dict):
            self._config = {}
        if self._config.get('ordinate') is not None:
            self._config["ordinate"] = self.line.rulerByName(self._config["ordinate"])
        self.checkGraphConfig()

        self._markdown = info.get("markdown", '')
        try:
            self._version = info['version']
        except KeyError:
            pass

        fp.close()
        self.setFullCheciMap()
        self.setSingleCheciMap()

    def version(self) -> str:
        return self._version

    def setVersion(self, v: str):
        self._version = v

    def addTrain(self, train: Train):
        self._trains.append(train)
        self.fullCheciMap[train.fullCheci()] = train
        self.addSingleCheciMap(train)

    def delTrain(self, train: Train):
        # import traceback
        try:
            if train.carriageCircuit() is not None:
                train.carriageCircuit().removeTrain(train)
            self._trains.remove(train)
            del self.fullCheciMap[train.fullCheci()]
            self.delSingleCheciMap(train)
        except Exception as e:
            print("del train: No such train!", train.fullCheci())
            # traceback.print_exc()

    def trains(self):
        for train in self._trains:
            yield train

    def show(self):
        self.line.show()
        # print(self.circuits)
        for train in self._trains:
            train.show()

    def setLine(self, line: Line):
        self.line = line

    def lineLength(self):
        try:
            return self.line.stations[-1]["licheng"]
        except IndexError:
            return 0

    def stations(self, reverse=False):
        if not reverse:
            for station in self.line.stations:
                yield station["zhanming"]
        else:
            for station in reversed(self.line.stations):
                yield station["zhanming"]

    def businessStationNames(self, passenger_only: bool, freight_only: bool):
        """
        2.1.1版本新增加。对办客和办货做筛选，由intervalCount过程调用。
        由于stations()函数调用范围太广，怕引起其他问题，故新增本函数。
        """
        for st in self.line.stations:
            if ((not passenger_only) or st.get("passenger", True)) and \
                    ((not freight_only) or st.get("freight", True)):
                yield st['zhanming']

    def stationDicts(self, reverse=False):
        if not reverse:
            for station in self.line.stations:
                yield station
        else:
            for station in reversed(self.line.stations):
                yield station

    def save(self, filename: str):
        """
        保存运行图文件
        """
        graph = {
            "line": self.line.outInfo(),
            "trains": [],
            "circuits": [],
            "config": self._config,
            "version": self._version,
        }
        for c in self._circuits:
            graph['circuits'].append(c.outInfo())
        try:
            graph["markdown"] = self._markdown
        except AttributeError:
            self._markdown = ''
            graph["markdown"] = ''
        if graph["config"]["ordinate"] is not None:
            if isinstance(graph["config"]["ordinate"], str):
                pass
            else:
                graph["config"]["ordinate"] = graph["config"]["ordinate"].name()

        for train in self._trains:
            graph["trains"].append(train.outInfo())

        with open(filename, 'w', encoding='utf8', errors='ignore') as fp:
            # print(graph["line"]["rulers"])
            json.dump(graph, fp, ensure_ascii=False)

    def UIConfigData(self):
        return self._config

    def sysConfigData(self):
        """
        1.4版本新增函数，专用返回系统默认设置。
        """
        return self._sysConfig

    def line_station_mileages(self):
        """
        生成器，返回 dict["zhanming"],dict["licheng"]
        """
        for dict in self.line.stations:
            yield dict["zhanming"], dict["licheng"]

    def line_station_mile_levels(self):
        for dict in self.line.stations:
            yield dict["zhanming"], dict["licheng"], dict["dengji"]

    def addEmptyRuler(self, name: str, different: bool = False):
        ruler = Ruler(name=name, different=different, line=self.line)
        self.line.rulers.append(ruler)
        return ruler

    def addRuler(self, ruler: Ruler):
        self.line.rulers.append(ruler)

    def delRuler(self, ruler: Ruler):
        """
        返回被删除的标尺是否是排图标尺
        """
        try:
            self.line.rulers.remove(ruler)
        except ValueError:
            pass
        if self.ordinateRuler() is ruler:
            self.setOrdinateRuler(None)
            return True
        return False

    def setOrdinateRuler(self, ruler: Ruler):
        self._config["ordinate"] = ruler

    def ordinateRuler(self):
        """
        2019.05.09：有时候莫名其妙会是一个str对象，就暴力解决问题。
        """
        ruler = self._config.setdefault("ordinate", None)
        if isinstance(ruler, str):
            print("Graph::ordinateRuler: ruler is str", ruler)
            return self.line.rulerByName(ruler)
        return ruler

    def setStationYValue(self, name, y):
        """
        设置某个站的纵坐标值。2019.02.02移植到line类并删除线性算法。
        """
        self.line.setStationYValue(name, y)

    def stationYValue(self, name: str):
        st_dict = self.stationByDict(name)
        if st_dict is not None:
            return st_dict.get('y_value', -1)
        else:
            return -1

    def stationMileYValues(self):
        for dct in self.line.stations:
            yield dct['zhanming'], dct['licheng'], dct.get('y_value', None)

    def trainFromCheci(self, checi: str, full_only=False) -> Train:
        """
        根据车次查找Train对象。如果full_only，则仅根据全车次查找；否则返回单车次匹配的第一个结果。
        若不存在，返回None。
        2019.02.03删除线性算法。
        """
        t = self.fullCheciMap.get(checi, None)
        if t is not None:
            return t
        if not full_only:
            selected = self.singleCheciMap.get(checi, None)
            if selected is not None:
                return selected[0]
        return None

    def multiSearch(self, checi: str):
        """
        非严格搜索。线性算法。
        """
        selected = []
        for train in self.trains():
            if checi in train.fullCheci() or checi in train.downCheci() or checi in train.upCheci():
                selected.append(train)
        return selected

    def lineName(self):
        return self.line.name

    def setLineName(self, name: str):
        self.line.name = name

    def firstStation(self):
        if self.line.stations:
            return self.line.stations[0]["zhanming"]
        else:
            return None

    def lastStation(self):
        try:
            return self.line.stations[-1]["zhanming"]
        except IndexError:
            return None

    def stationInLine(self, name: str, strict=False):
        return self.line.stationInLine(name, strict)

    def setStationIsShow(self, name: str, show: bool):
        """
        不支持域解析符。2019.02.02删除线性算法。
        """
        st = self.line.stationDictByName(name, True)
        if st is None:
            raise Exception("setStationIsShow: no such station")
        else:
            st['show'] = show

    def stationIsShow(self, name: str):
        dct = self.line.stationDictByName(name)
        if dct is None:
            raise Exception("No such station")
        return dct.setdefault('show', True)

    def graphFileName(self):
        return self.filename

    def setGraphFileName(self, filename: str):
        self.filename = filename

    def isEmpty(self):
        if self.line.stations:
            return False
        else:
            return True

    def stationCount(self):
        return len(self.line.stations)

    def clearLineStationInfo(self):
        del self.line.stations
        self.line.stations = []

    def stationExisted(self, name: str):
        self.line.stationExisted(name)

    def addStationDict(self, info: dict):
        self.line.stations.append(info)

    def adjustLichengTo0(self):
        if self.isEmpty():
            return
        start_mile = self.line.stations[0]["licheng"]
        for st in self.line.stations:
            st["licheng"] = st["licheng"] - start_mile

    def trainCount(self):
        return len(self._trains)

    def rulers(self):
        for ruler in self.line.rulers:
            yield ruler

    def gapBetween(self, st1: str, st2: str):
        """
        计算两个站间距离
        """
        station1 = self.stationByDict(st1)
        station2 = self.stationByDict(st2)

        if station1 is None or station2 is None:
            raise Exception("No such station", st1, st2)

        return abs(station1["licheng"] - station2["licheng"])

    def lineSplited(self):
        """
        返回本线是否存在上下行分设站的情况
        :return:
        """
        if self.line.isSplited():
            return True
        return False

    def rulerNameExisted(self, name, ignore: Ruler = None):
        for r in self.line.rulers:
            if r is not ignore and r.name() == name:
                return True
        return False

    def isNewRuler(self, ruler: Ruler):
        for r in self.line.rulers:
            if ruler is r:
                return False
        return True

    def stationDirection(self, name: str):
        return self.line.stationViaDirection(name)

    def lineStationBusiness(self, name: str, passenger: int, default=False) -> bool:
        """
        2.0.2新增，返回车站是否办理业务。passenger是Train中规定的枚举常量，标志是否办客。
        如果找不到，返回default。
        """
        dct = self.line.stationDictByName(name)
        if dct is None:
            # print("graph::lineStationBusiness: no such station! return",default,name)
            return default

        if passenger == Train.PassengerTrue:
            # print("graph::lineStationBusiness passengerTrue",dct.get('passenger',"无数据"))
            return dct.get('passenger', True)
        else:
            # print("graph::lineStationBusiness passengerFalse",dct.get("freight","无数据"))
            return dct.get("freight", True)

    def formerBothStation(self, name: str):
        """
        寻找本站往前的第一个【下行方向通过】的站。
        TODO 2019.02.02 保留线性算法。下一个函数同。
        """
        former_dict = None
        for st in self.line.stations:
            if st["zhanming"] == name:
                return former_dict

            if st["direction"] == 0x3:
                former_dict = st
        raise Exception("No former station")

    def latterBothStation(self, name: str):
        start = False
        for st in self.line.stations:
            if st["zhanming"] == name:
                start = True
            if start and st["direction"] == 0x3:
                return st
        raise Exception("No latter station")

    def stationLevel(self, name: str):
        """
        返回车站等级。若不存在，返回None；若没有这个字段，设为并返回4. 不支持域解析符
        """
        st = self.line.stationDictByName(name, strict=True)
        if st is None:
            return None
        return st.setdefault('dengji', 4)

    def setNotShowTypes(self, not_show):
        self.UIConfigData()["not_show_types"] = not_show
        for train in self.trains():
            if train.type in not_show:
                train.setIsShow(False, affect_item=False)
            else:
                train.setIsShow(True, affect_item=False)

    def setDirShow(self, down, show):
        """
        约定此处指入图时的上下行
        """
        for train in self.trains():
            if train.firstDown() == down and train.type not in self.UIConfigData()['not_show_types']:
                train.setIsShow(show, affect_item=False)

    def trainExisted(self, train: Train, ignore: Train = None):
        """
        比较Train对象，线性算法
        """
        for t in self._trains:
            if train is t and t is not ignore:
                return True
        return False

    def checiExisted(self, checi: str, ignore: Train = None):
        """
        比较全车次。2019.02.03替换掉线性算法。
        """
        # for t in self._trains:
        #     if t is not ignore and t.fullCheci() == checi:
        #         return True
        # return False
        t = self.fullCheciMap.get(checi, None)
        if t is not None and (ignore is None or t is not ignore):
            return True
        return False

    def rulerCount(self):
        return len(self.line.rulers)

    def checiType(self, checi: str) -> str:
        """
        2.0.2新增。根据系统设置的判断规则，返回车次对应的类型。如果不符合任何一个，返回 其他。
        """
        for nm, rg, _ in self.UIConfigData()['type_regex']:
            try:
                rgx = re.compile(rg)
            except:
                print("Invalid Regex! ",rg)
                continue
            if re.match(rgx, checi):
                return nm
        return '其他'

    def checiTypePassenger(self, checi: str) -> (str, int):
        """
        根据车次返回类型以及是否是客车。是否是客车按照Train中定义的常量。
        如果不符合任何一个，返回 其他, PassengerAuto。
        """
        for nm, rg, ps in self.UIConfigData()['type_regex']:
            if re.match(rg, checi):
                if ps:
                    return nm, Train.PassengerTrue
                else:
                    return nm, Train.PassengerFalse
        return '其他', Train.PassengerAuto

    def typePassenger(self, tp: str, default=Train.PassengerAuto) -> int:
        """
        根据类型返回是否为客车。返回是Train中的PassengerTrue或PassengerFalse，如果找不到返回默认。
        """
        for name, _, ps in self.UIConfigData()['type_regex']:
            if name == tp:
                if ps:
                    return Train.PassengerTrue
                else:
                    return Train.PassengerFalse
        return default

    def stationTimeTable(self, name: str):
        """
        返回车站的图定时刻表
        list<dict>
        dict{
            "station_name":str,
            "ddsj":datetime,
            "cfsj":datetime,
            "down":bool,
            "train":Train,
        }
        """
        timeTable = []
        for train in self.trains():
            st_dict = train.stationDict(name)
            if st_dict is None:
                continue
            else:
                node = {
                    "ddsj": st_dict["ddsj"],
                    "cfsj": st_dict["cfsj"],
                    "station_name": st_dict["zhanming"],
                    "down": train.stationDown(st_dict['zhanming'], self),
                    "note": st_dict.get("note", ''),
                    "train": train,
                }
                timeTable.append(node)

        # 排序
        for i in range(len(timeTable) - 1):
            t = i
            for j in range(i + 1, len(timeTable)):
                if timeTable[j]["ddsj"] < timeTable[t]["ddsj"]:
                    t = j
            temp = timeTable[t];
            timeTable[t] = timeTable[i];
            timeTable[i] = temp
        return timeTable

    def reverse(self):
        """
        反排运行图
        """
        length = self.lineLength()

        # 里程调整
        for st_dict in self.line.stations:
            st_dict["licheng"] = length - st_dict["licheng"]
        self.line.stations.reverse()

        # 列车上下行调整、上下行车次交换
        for train in self._trains:
            # 上下行交换
            train.reverseAllItemDown()

            # 车次交换
            temp = train.setCheci(train.fullCheci(), train.upCheci(), train.downCheci())

    def downTrainCount(self):
        count = 0
        for train in self.trains():
            if train.firstDown() is True:
                count += 1
        return count

    def upTrainCount(self):
        count = 0
        for train in self.trains():
            if train.firstDown() is False:
                count += 1
        return count

    def loadTrcGraph(self, filename):
        """
        阅读旧版trc格式的运行图
        """
        fp = open(filename, encoding='utf-8', errors='ignore')
        self.line.forbid.setDifferent(False)
        self.line.forbid.setShow(True, True)
        inTrainArea = False
        now_list = []
        last_name = None
        circuit_dict = {}
        for i, line in enumerate(fp):
            line = line.strip()
            if not line:
                continue
            if not inTrainArea and line == "===Train===":
                inTrainArea = True

            if line[0] == '-':
                break

            # 处理线路信息部分
            if not inTrainArea:
                if line == "***Circuit***":
                    continue
                elif i == 1:
                    self.setLineName(line)
                else:
                    # 线路信息部分
                    try:
                        splited = line.split(',')
                        st_name = splited[0]
                        self.line.addStation_by_info(splited[0], int(splited[1]), int(splited[2]))
                        if last_name is not None:
                            try:
                                start_str, end_str = splited[9].split('-', 1)
                                begin = datetime.strptime(start_str, '%H:%M')
                                end = datetime.strptime(end_str, '%H:%M')
                                self.line.forbid.addForbid(last_name, st_name, begin, end)
                            except Exception as e:
                                pass
                        last_name = st_name
                    except:
                        pass

            # 处理列车信息部分
            else:
                # 这部分从trc_check_new中复制过来
                if line != '===Train===':
                    now_list.append(line)
                else:
                    self._decodeTrcTrain(now_list, circuit_dict)
                    now_list = []
        self._decodeTrcTrain(now_list, circuit_dict)
        self._decodeTrcCircuit(circuit_dict)
        self.setGraphFileName('')

    def _decodeTrcTrain(self, now_list: list, circuit_dict: dict):
        """
        阅读trc中单个车次的信息，不含===Train===标志头。
        circuit_dict: 抽取车次中含有的交路信息。数据结构为
        Dict<List<Tuple<int,Train>>>
        eg.
        {
            "CRH380-2081":[
                (0, Train<G1>),
                (1, Train<G4>),
                ...
            ],
            ...
        }
        交路数据保证只有一个下划线“_”，且split之后是一个整数。
        """
        train = Train(self)
        for i, line in enumerate(now_list):
            if i == 0:
                splited = line.split(',')
                train.setCheci(splited[1], splited[2], splited[3])
                if len(splited) >= 5:
                    circuit_str = splited[4]
                    try:
                        num = int(circuit_str.split('_')[-1])
                        name = circuit_str.split('_')[0]
                    except ValueError:
                        if circuit_str not in ('NA', ''):
                            print("Graph::decodeTrcTrain: Unexpected circuit info:", circuit_str)
                    else:
                        circuit_dict.setdefault(name, []).append((num, train))


            elif i == 1:
                train.setStartEnd(sfz=line)
            elif i == 2:
                train.setStartEnd(zdz=line)
            else:
                splited = line.split(',')
                train.addStation(splited[0], splited[1], splited[2])
        train.autoTrainType()
        if train.timetable:
            self.addTrain(train)

    def _decodeTrcCircuit(self, circuit_dict: dict):
        """
        解析从前面收集到的交路数据，生成交路对象。
        """
        for name, lst in circuit_dict.items():
            lst.sort()
            circuit = Circuit(self, name)
            for _, train in lst:
                circuit.addTrain(train)
            self.addCircuit(circuit)

    def jointGraph(self, graph, former: bool, reverse: bool, line_only: bool):
        """
        拼接两运行图。
        :param graph: 另一运行图
        :param former: 另一运行图是否在本运行图前侧链接
        :param reverse: 另一运行图是否转置
        """
        if reverse:
            graph.reverse()

        if not line_only:
            # 车次连接
            for train_append in graph.trains():
                if self.checiExisted(train_append.fullCheci()):
                    # 本线有相同车次
                    train_main: Train = self.trainFromCheci(train_append.fullCheci())
                    train_main.delNonLocal(self)
                    train_append.delNonLocal(graph)
                    # 方向以本线为准
                    # down表示列车在两线路连接点附近的上下行情况。
                    if former:
                        down = train_main.firstDown()
                    else:
                        down = train_main.lastDown()
                    if down is None:
                        # 如果本线无法判断，例如Z90终到石家庄在京石段只有一个站，则用另一条线的。
                        if former:
                            down = train_append.lastDown()
                        else:
                            down = train_append.firstDown()
                    if down is None:
                        # 如果都无法判断，直接判断为下行车次
                        print("cannot judge down. use default.", train_main.fullCheci())
                        down = True

                    train_former = not (down ^ former)
                    train_main.jointTrain(train_append, train_former, graph)  # 当前站点已经拼接好

                else:
                    self.addTrain(train_append)

        # 线路连接
        if former:
            for station in self.stationDicts():
                station["licheng"] += graph.lineLength()

            for st in reversed(graph.line.stations):
                if not self.stationExisted(st["zhanming"]):
                    self.line.addStation_by_origin(st, index=0)
        else:
            length = self.lineLength()
            for st in graph.stationDicts():
                st["licheng"] += length
                if not self.stationExisted(st["zhanming"]):
                    self.line.addStation_by_origin(st)

        # 标尺的处理，直接复制过来就好
        for ruler in graph.rulers():
            thisruler = self.line.rulerByName(ruler.name())
            if thisruler is not None:
                thisruler._nodes += ruler._nodes
            else:
                self.addRuler(ruler)

    def resetAllItems(self):
        for train in self.trains():
            train.setItem(None)

    def stationMile(self, name: str):
        """
        返回车站的里程数据，若不存在返回-1.不支持域解析符。2019.02.03删除线性算法。
        2019.02.23改为：支持域解析符。
        """
        st = self.line.stationDictByName(name)
        if st is None:
            return -1
        return st["licheng"]

    def adjacentStation(self, name: str, ignore: list):
        index = self.stationIndex(name)
        if index > 0:
            if self.line.stations[index - 1]['zhanming'] not in ignore:
                # 2019.02.23修改，条件少了zhanming，not in的判断相当于没用
                return self.line.stations[index - 1]["zhanming"]

        if index < len(self.line.stations) - 1:
            if self.line.stations[index + 1]['zhanming'] not in ignore:
                return self.line.stations[index + 1]["zhanming"]
        print("no adj")
        return None

    def stationIndex(self, name: str):
        for i, st in enumerate(self.line.stations):
            if stationEqual(st["zhanming"], name):
                return i
        raise Exception("No such station", name)

    def stationByDict(self, name: str, strict=False):
        """
        根据站名返回dict对象，函数名写错了。支持域解析符。
        2019.02.02删除线性算法。
        """
        return self.line.stationDictByName(name, strict)

    def passedStationCount(self, st1: str, st2: str, down: bool) -> int:
        """
        检查以st1为发站，st2为到站，方向为down的区间内有多少个站。2.0新增。
        """
        s1 = self.stationIndex(st1)
        s2 = self.stationIndex(st2)
        dir_ = 0b1 if down else 0b10
        cnt = 0
        t1 = min((s1, s2))
        t2 = max((s1, s2))
        # print("t1 t2",t1,t2)
        for i in range(t1 + 1, t2):
            dct = self.line.stationDictByIndex(i)
            if dir_ & dct.get('direction', 0b11):
                cnt += 1
        # print("passedStationCount",st1,st2,down,cnt)
        return cnt

    def resetStationName(self, old, new, auto_field=False):
        old_dict = self.stationByDict(old)
        if old_dict is not None:
            old_dict["zhanming"] = new if not auto_field else new.split('::')[0]

        for ruler in self.line.rulers:
            ruler.changeStationName(old, new)
        for train in self.trains():
            if train.isSfz(old):
                train.sfz = new
            if train.isZdz(old):
                train.zdz = new
            if train._localFirst == old:
                train._localFirst = new
            elif train._localLast == old:
                train._localLast = new
            st_dict = train.stationDict(old)
            if st_dict is not None:
                st_dict["zhanming"] = new
        self.line.changeStationNameUpdateMap(old, new)

    def addTrainByGraph(self, graph, cover=False):
        """
        添加车次，返回数量
        """
        num = 0
        for train in graph.trains():
            if train.localCount(self) >= 2:
                if not self.checiExisted(train.fullCheci()):
                    num += 1
                    self.addTrain(train)
                elif cover:
                    num += 1
                    t = self.trainFromCheci(train.fullCheci())
                    self.delTrain(t)
                    self.addTrain(train)

        return num

    def setMarkdown(self, mark: str):
        self._markdown = mark

    def markdown(self):
        try:
            return self._markdown
        except AttributeError:
            self._markdown = ""
            return ''

    def save_excel(self, filename: str):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = f'{self.firstStation()}-{self.lastStation()}间列车时刻表'

        # 写入左侧表头
        ws['A3'] = '始发站'
        ws.merge_cells('A3:A4')
        ws['A5'] = '终到站'
        ws.merge_cells('A5:A6')
        ws['A7'] = '列车种类'
        ws.merge_cells('A7:A8')
        ws['A9'] = '车次'
        ws['A10'] = '车站'
        for row in range(3, 9):
            ws.row_dimensions[row].font = Font(name='SimSum', size=9)
            ws.row_dimensions[row].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 9.7

        start = 11  # 从第11行开始表格
        # 写入车站
        station_row_dict = {}
        cur = 11
        for station in self.stations():
            ws.cell(row=cur, column=1, value=station)
            ws.merge_cells(start_row=cur, end_row=cur + 1, start_column=1, end_column=1)
            station_row_dict[station] = cur
            ws.row_dimensions[cur].height = 9.7
            ws.row_dimensions[cur + 1].height = 9.7
            ws.row_dimensions[cur].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[cur + 1].alignment = Alignment(horizontal='center', vertical='center')
            cur += 2
        ws.column_dimensions['A'].width = 12

        # 写入车次，先下行
        last_merge_sfz, last_merge_zdz, last_merge_type = 1, 1, 1
        col = 2
        last_train = None
        for train in self.trains():
            for dct in train.itemInfo():
                if not dct['down']:
                    continue
                if last_train and train.sfz == last_train.sfz:
                    try:
                        ws.unmerge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col)
                else:
                    ws.merge_cells(start_row=3, end_row=4, start_column=col, end_column=col)
                    last_merge_sfz = col
                ws.cell(row=3, column=col, value=train.sfz)

                if last_train and train.zdz == last_train.zdz:
                    try:
                        ws.unmerge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col)
                else:
                    ws.merge_cells(start_row=5, end_row=6, start_column=col, end_column=col)
                    last_merge_zdz = col
                c = ws.cell(row=5, column=col, value=train.zdz)
                col_str = c.column
                ws.column_dimensions[col_str].width = 6  # 设置列宽为5

                if last_train and train.type == last_train.type:
                    try:
                        ws.unmerge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col)
                else:
                    ws.merge_cells(start_row=7, end_row=8, start_column=col, end_column=col)
                    last_merge_type = col
                ws.cell(row=7, column=col, value=train.type)

                checi = train.fullCheci()
                if '/' in checi:
                    ws.cell(row=9, column=col, value=checi.split('/')[0])
                    ws.cell(row=10, column=col, value='/' + checi.split('/', maxsplit=1)[1])
                else:
                    ws.cell(row=9, column=col, value=checi)
                    ws.merge_cells(start_row=9, end_row=10, start_column=col, end_column=col)

                last_dict = None

                # 时刻表循环
                for st_dict in train.timetable:
                    for i, s in station_row_dict.items():
                        if stationEqual(i, st_dict['zhanming']):
                            row = s
                            break
                    else:
                        continue

                    if train.isSfz(st_dict['zhanming']):
                        ws.cell(row=row, column=col, value='')
                        ws.cell(row=row + 1, column=col, value=self.outTime(st_dict['cfsj'], True))

                    elif train.isZdz(st_dict["zhanming"]):
                        ws.cell(row=row, column=col, value=self.outTime(st_dict['ddsj'], True))
                        ws.cell(row=row + 1, column=col, value='    --')

                    elif train.stationStopped(st_dict):
                        # 本站停车，无条件写入完整到达时刻和不完整出发时刻
                        ddsj_str = f'{st_dict["ddsj"].hour:2d}:{st_dict["ddsj"].minute:02d}'
                        sec = st_dict['ddsj'].second
                        if sec:
                            ddsj_str += f"{sec:02d}"
                        else:
                            ddsj_str += '  '
                        ws.cell(row=row, column=col, value=ddsj_str)
                        if st_dict['ddsj'].hour == st_dict['cfsj'].hour:
                            cfsj_str = '   '
                        else:
                            cfsj_str = f"{st_dict['cfsj'].hour:2d}:"
                        cfsj_str += f'{st_dict["cfsj"].minute:02d}'
                        sec = st_dict['cfsj'].second
                        if sec:
                            ddsj_str += f"{sec:02d}"
                        else:
                            ddsj_str += '  '
                        ws.cell(row=row + 1, column=col, value=cfsj_str)

                    else:
                        give_hour = False
                        if not last_dict:
                            give_hour = True
                        elif last_dict['cfsj'].hour != st_dict['ddsj'].hour:
                            give_hour = True
                        ws.cell(row=row, column=col, value='   ...')
                        tgsj_str = f'{st_dict["ddsj"].hour:2d}:' if give_hour else '   '
                        tgsj_str += f'{st_dict["ddsj"].minute:02d}'
                        sec = st_dict['ddsj'].second
                        if sec:
                            tgsj_str += f"{sec:02d}"
                        else:
                            tgsj_str += '  '
                        ws.cell(row=row + 1, column=col, value=tgsj_str)
                    last_dict = st_dict
                col += 1
                last_train = train

        # 上行
        for train in self.trains():
            for dct in train.itemInfo():
                if dct['down']:
                    continue
                if last_train and train.sfz == last_train.sfz:
                    try:
                        ws.unmerge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col)
                else:
                    ws.merge_cells(start_row=3, end_row=4, start_column=col, end_column=col)
                    last_merge_sfz = col
                c = ws.cell(row=3, column=col, value=train.sfz)
                col_str = c.column
                ws.column_dimensions[col_str].width = 6  # 设置列宽为5

                if last_train and train.zdz == last_train.zdz:
                    try:
                        ws.unmerge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col)
                else:
                    ws.merge_cells(start_row=5, end_row=6, start_column=col, end_column=col)
                    last_merge_zdz = col
                ws.cell(row=5, column=col, value=train.zdz)

                if last_train and train.type == last_train.type:
                    try:
                        ws.unmerge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col)
                else:
                    ws.merge_cells(start_row=7, end_row=8, start_column=col, end_column=col)
                    last_merge_type = col
                ws.cell(row=7, column=col, value=train.type)

                checi = train.fullCheci()
                if '/' in checi:
                    ws.cell(row=9, column=col, value=checi.split('/')[0])
                    ws.cell(row=10, column=col, value='/' + checi.split('/', maxsplit=1)[1])
                else:
                    ws.cell(row=9, column=col, value=checi)
                    ws.merge_cells(start_row=9, end_row=10, start_column=col, end_column=col)

                last_dict = None
                # 时刻表循环
                for st_dict in train.timetable:
                    for i, s in station_row_dict.items():
                        if stationEqual(i, st_dict['zhanming']):
                            row = s
                            break
                    else:
                        continue

                    if train.isSfz(st_dict['zhanming']):
                        ws.cell(row=row + 1, column=col, value='')
                        ws.cell(row=row, column=col, value=self.outTime(st_dict['cfsj'], True))

                    elif train.isZdz(st_dict["zhanming"]):
                        ws.cell(row=row + 1, column=col, value=self.outTime(st_dict['ddsj'], True))
                        ws.cell(row=row, column=col, value='    --')

                    elif train.stationStopped(st_dict):
                        # 本站停车，无条件写入完整到达时刻和不完整出发时刻
                        ddsj_str = f'{st_dict["ddsj"].hour:2d}:{st_dict["ddsj"].minute:02d}'
                        sec = st_dict['ddsj'].second
                        if sec:
                            ddsj_str += f"{sec:02d}"
                        else:
                            ddsj_str += '  '
                        ws.cell(row=row + 1, column=col, value=ddsj_str)
                        if st_dict['ddsj'].hour == st_dict['cfsj'].hour:
                            cfsj_str = '   '
                        else:
                            cfsj_str = f"{st_dict['cfsj'].hour:2d}:"
                        cfsj_str += f'{st_dict["cfsj"].minute:02d}'
                        sec = st_dict['cfsj'].second
                        if sec:
                            ddsj_str += f"{sec:02d}"
                        else:
                            ddsj_str += '  '
                        ws.cell(row=row, column=col, value=cfsj_str)

                    else:
                        give_hour = False
                        if not last_dict:
                            give_hour = True
                        elif last_dict['cfsj'].hour != st_dict['ddsj'].hour:
                            give_hour = True
                        ws.cell(row=row + 1, column=col, value='   ...')
                        tgsj_str = f'{st_dict["ddsj"].hour:2d}:' if give_hour else '   '
                        tgsj_str += f'{st_dict["ddsj"].minute:02d}'
                        sec = st_dict['ddsj'].second
                        if sec:
                            tgsj_str += f"{sec:02d}"
                        else:
                            tgsj_str += '  '
                        ws.cell(row=row, column=col, value=tgsj_str)
                col += 1
                last_train = train

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center',
                                                                   vertical='center', shrink_to_fit=True)
                ws.cell(row=row, column=col).font = Font(name='宋体', size=9)

        wb.save(filename)

    def outTime(self, tgsj, give_hour: bool):
        tgsj_str = f'{tgsj.hour:2d}:' if give_hour else '   '
        tgsj_str += f'{tgsj.minute:02d}'
        sec = tgsj.second
        if sec:
            tgsj_str += f"{sec:02d}"
        else:
            tgsj_str += '  '
        return tgsj_str

    def getIntervalTrains(self, start, end, trainFilter,*,businessOnly=False,stoppedOnly=False):
        """
        返回某个区间办客车次列表。数据结构为list<dict>。
        //2.1版本修改逻辑为：两站皆办理业务才被选入。
        2019.06.29修改逻辑：选入的条件由输入参数给定。其中stoppedOnly包含始发终到情况。
        dict{
            'train':train object,
            'isSfz':boolean,
            'isZdz':boolean,
            'from':str,
            'to':str,
        """
        interval_list = []
        for train in self.trains():
            if not trainFilter.check(train):
                continue
            # b1 = train.stationStopBehaviour(start)
            # b2 = train.stationStopBehaviour(end)
            start_dict = train.stationDict(start)
            end_dict = train.stationDict(end)
            # p1 = train.stationBusiness(start_dict)
            # p2 = train.stationBusiness(end_dict)
            if not train.stationBefore(start, end):
                continue
            if not (self.judgeStopAndBusiness(train,start_dict,businessOnly,stoppedOnly) and
                    self.judgeStopAndBusiness(train,end_dict,businessOnly,stoppedOnly)):
                continue
            isSfz = train.isSfz(start)
            isZdz = train.isZdz(end)
            train_dict = {
                'train': train,
                'isSfz': isSfz,
                'isZdz': isZdz,
                'from': start_dict['zhanming'],
                'to': end_dict['zhanming']
            }
            interval_list.append(train_dict)
        return interval_list

    def judgeStopAndBusiness(self,train:Train,dct:dict,bOnly:bool,sOnly:bool):
        """
        为上一个函数服务的工具性函数。判断时刻表中某车站是否符合对营业和停车的要求。
        表达式写的丑是为了利用短路性提高效率。
        """
        zm = dct['zhanming']
        return (not sOnly or train.stationStopped(dct) or train.isSfz(zm) or train.isZdz(zm)) and\
               (not bOnly or train.stationBusiness(dct))

    def getIntervalCount(self, fromOrTo, isStart, trainFilter, passenger_only=False, freight_only=False):
        """
        获取区间对数表。
        :param fromOrTo:发站或到站
        :param isStart: True for start station, vice versa
        后两个参数：是否仅包括办客和办货的车站。中心站（fromOrTo）不受限制。
        返回数据结构list<dict>
        dict{
        'from'
        """
        infoList = []
        if isStart:
            for st in self.businessStationNames(passenger_only, freight_only):
                if not stationEqual(fromOrTo, st):
                    infoList.append({'from': fromOrTo, 'to': st, 'info': self.getIntervalTrains(fromOrTo, st,
                                                                                                trainFilter)})
        else:
            for st in self.businessStationNames(passenger_only, freight_only):
                if not stationEqual(fromOrTo, st):
                    infoList.append({'to': fromOrTo, 'from': st, 'info': self.getIntervalTrains(st, fromOrTo,
                                                                                                trainFilter)})

        count_list = []
        for info_dict in infoList:
            info = info_dict['info']
            count = len(tuple(info))
            countSfz = len([1 for st in info if st['isSfz']])
            countZdz = len([1 for st in info if st['isZdz']])
            countSfZd = len([1 for st in info if st['isZdz'] and st['isSfz']])
            int_dict = {
                'from': info_dict['from'],
                'to': info_dict['to'],
                'count': count,
                'countSfz': countSfz,
                'countZdz': countZdz,
                'countSfZd': countSfZd
            }
            count_list.append(int_dict)
        return count_list

    def stationByIndex(self, idx):
        return self.line.stations[idx]

    def resetAllTrainsLocalFirstLast(self):
        """
        当线路数据更新时，重置所有列车的localFirst/Last。
        """
        for train in self.trains():
            train.updateLocalFirst(self)
            train.updateLocalLast(self)

    def setMargin(self, ruler_label, mile_label, station_label, left_and_right, top_and_bottom, system=False) -> bool:
        """
        1.4版本修改
        用户设置页边距。注意参数含义与本系统内部使用的不同。返回是否变化。
        """
        left_white = 15
        right_white = 10
        margins = {
            "left_white": left_white,
            "right_white": right_white,
            "left": ruler_label + mile_label + station_label + left_and_right + left_white,
            "up": top_and_bottom,
            "down": top_and_bottom,
            "right": left_and_right + station_label + right_white,
            "label_width": station_label,
            "mile_label_width": mile_label,
            "ruler_label_width": ruler_label,
        }
        UIDict = self.UIConfigData() if not system else self.sysConfigData()
        changed = UIDict.get('margins', None) != margins
        UIDict["margins"] = margins
        return changed

    def toTrc(self, filename):
        fp = open(filename, 'w', encoding='utf-8', errors='ignore')
        fp.write('***Circuit***\n')
        fp.write(f"{self.lineName() if self.lineName() else '列车运行图'}\n")
        fp.write(f"{int(self.lineLength())}\n")
        last_dct = None
        for dct in self.stationDicts():
            fp.write(f"{dct['zhanming']},{int(dct['licheng'])},{dct['dengji']},"
                     f"{'false' if dct.get('show',True) else 'true'},,true,4,1440,1440,")
            if last_dct is not None:
                node = self.line.forbid.getInfo(last_dct['zhanming'], dct['zhanming'])
                if node is not None and node['begin'] != node['end']:
                    fp.write(f"{node['begin'].strftime('%H:%M')}-{node['end'].strftime('%H:%M')}\n")
                else:
                    fp.write('\n')
            else:
                fp.write('\n')
            last_dct = dct
        for train in self.trains():
            fp.write('===Train===\n')
            fp.write(f"trf2,{train.fullCheci()},{train.downCheci()},{train.upCheci()}")
            circuit = train.carriageCircuit()
            if isinstance(circuit, Circuit):
                fp.write(f",{circuit.name().replace('_','-')}_{circuit.trainOrderNum(train)}\n")
            else:
                fp.write(',NA\n')
            fp.write(f"{train.sfz if train.sfz else 'null'}\n")
            fp.write(f"{train.zdz if train.zdz else 'null'}\n")
            for name, ddsj, cfsj in train.station_infos():
                fp.write(f"{name},{ddsj.strftime('%H:%M:%S')},{cfsj.strftime('%H:%M:%S')},true,NA\n")
        fp.write('---Color---\n')
        for train in self.trains():
            color_str = train.color(self)
            fp.write(f"{train.fullCheci()},{int(color_str[1:3],base=16)},"
                     f"{int(color_str[3:5],base=16)},{int(color_str[5:7],base=16)}\n")
        fp.write('---LineType---\n')
        for train in self.trains():
            if train.lineWidth() != 0:
                fp.write(f"{train.fullCheci()},0,{train.lineWidth()}")
        fp.close()

    def circuitByName(self, name: str, *, throwError=True) -> Circuit:
        """
        根据交路名查找交路对象。如果没有对应交路，抛出CircuitNotFoundError。
        """
        for circuit in self._circuits:
            if circuit.name() == name:
                return circuit
        if throwError:
            raise CircuitNotFoundError(name)
        else:
            return None

    def circuits(self):
        for c in self._circuits:
            yield c

    def circuitCount(self):
        return len(self._circuits)

    def addCircuit(self, circuit: Circuit):
        """
        若名称已存在，抛出CircuitExistedError
        """
        if self.circuitByName(circuit.name(), throwError=False) is not None:
            raise CircuitExistedError(circuit.name())
        self._circuits.append(circuit)

    def delCircuit(self, circuit: Circuit):
        """
        若不存在，抛出CircuitNotFoundError
        """
        assert isinstance(circuit, Circuit)
        try:
            self._circuits.remove(circuit)
        except ValueError:
            raise CircuitNotFoundError(circuit.name())
        for node in circuit.nodes():
            try:
                node.train().setCarriageCircuit(None)
            except AttributeError:
                print("Graph::delCircuit: Unexcpeted node.train", node)

    def checkGraph(self)->str:
        """
        对运行图文件可能出现的问题做启动时的检查。返回是报错的字符串。没有其他影响。
        如果没有问题，返回空串。
        """
        report = ""
        # 检查正则表达式的问题
        i = 0
        found = False
        for name,rg,_ in self.UIConfigData()['type_regex']:
            try:
                re.compile(rg)
            except:
                if rg == r'7\d+{3}':
                    self.UIConfigData()['type_regex'][i] = (name,r'7\d{3}',_)
                else:
                    found=True
                    report += f"列车类型[{name}]的正则表达式[{rg}]错误！\n"
            else:
                pass
            i+=1
        if found:
            report += "请至“运行图设置”（ctrl+G）面板中“类型管理”中修正错误的正则表达式。\n"

        if report:
            report += "此信息提示当前运行图文件可能不符合此版本软件的要求。您可以忽略此提示，但有可能导致" \
                      "程序运行异常。下次打开本文件时，如果仍未解决，此消息仍会显示。\n"
        return report

    def modelList(self)->list:
        """
        2019.06.25新增。返回所有不同的车底类型列表。
        目前只由trainFilter中筛选的有关功能调用，简化起见，不作为属性保存，每次都遍历得到。
        """
        lst = []
        for circuit in self.circuits():
            lst.append(circuit.model())
        return list(set(lst))

    def ownerList(self)->list:
        """
        返回所有不同的担当局段表。
        """
        lst = []
        for circuit in self.circuits():
            lst.append(circuit.owner())
        return list(set(lst))

    def clearAll(self):
        """
        清空所有数据
        """
        self.clearLineStationInfo()
        self._trains = []
        self._circuits = []
        self.typeList = []
        self.fullCheciMap = {}
        self.singleCheciMap = {}
        self._markdown = ''


if __name__ == '__main__':
    graph = Graph()
    graph.loadGraph("source/output.json")
    graph.show()
    graph.save("source/test.json")
