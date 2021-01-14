"""
运行图类.
同时管理self._config和self._sysConfig两套设置系统，两套系统完全分离。
"""
from .line import Line,LineStation
from .ruler import Ruler
from .train import Train, TrainStation
from .circuit import Circuit, CircuitNode
from copy import copy
from Timetable_new.utility import stationEqual, strToTime
import json, re
from datetime import datetime
from ..pyETRCExceptions import *
from typing import List, Union, Tuple, Dict
from enum import Enum
from functools import reduce
from math import sqrt
import time

config_file = 'config.json'

import cgitb

cgitb.enable(format='text')


class Graph:
    """
    运行图类，数据结构：
    Line line;
    List<Train> _trains=[];
    List<Circuit> _circuit;//交路数据
    Dict _config;//系统设置，主要是UI
    str markdown;//附注
    """

    def __init__(self):
        """
        构造空类，不考虑读文件
        """
        self.filename = ""
        self._version = ""
        self._sysVersion = ""
        self.line = Line()
        self._trains = []
        self._circuits = []
        self._config = None
        self._sysConfig = None
        self.readSysConfig()  # 初始化并校验系统默认配置文件
        self.typeList = []  # public data
        self.initGraphConfig()
        self._markdown = ''
        self.fullCheciMap = {}  # 全车次查找表 str|->Train
        self.singleCheciMap = {}  # 单车次查找表str|->list<Train>

    def readSysConfig(self):
        """
        1.4新定义函数。读取并检查系统默认设置。
        """
        try:
            with open(config_file, encoding='utf-8', errors='ignore') as fp:
                self._sysConfig = json.load(fp)
        except:
            print("配置文件config.json加载失败，使用系统默认设置。")
            self._sysConfig = {}
        self.checkSysConfig()

    def checkSysConfig(self):
        """
        1.4版本新增函数。每次在系统设置中增加新的字段时，要修改本函数。
        """
        default_config = {
            "seconds_per_pix": 15.0,
            "seconds_per_pix_y": 8.0,
            "pixes_per_km": 4.0,
            "grid_color": "#AAAA7F",
            "text_color": "#0000FF",
            "default_keche_width": 1.5,
            "default_huoche_width": 0.75,
            "default_db_file": "linesNew.pyetlib",
            "start_hour": 0,
            "end_hour": 24,
            "minutes_per_vertical_line": 10.0,
            "bold_line_level": 2,
            "show_line_in_station": True,
            "start_label_height": 30,
            "end_label_height": 15,
            "table_row_height": 30,
            "link_line_height": 10,
            "show_time_mark": 1,  # 显示详细停点。0-不显示，1-仅显示选中车次，2-显示所有车次
            "max_passed_stations": 3,  # 至多跨越站数。超过这个数将被分成两段运行图。
            "avoid_cover":True,  # 自动偏移标签避免重叠
            "base_label_height":15,  # 基准标签高度，启用自动偏移时有效
            "step_label_height":20,  # 每一次自动偏移的标签高度
            "end_label_checi":False,  # 结束标签是否要显示车次
            "default_colors": {"快速": "#FF0000",
                               "特快": "#0000FF",
                               "直达特快": "#FF00FF",
                               "动车组": "#804000",
                               "动车": "#804000",
                               "高速": "#FF00BE",
                               "城际": "#FF33CC",
                               "default": "#008000"
                               },
            "margins": {
                "left_white": 15,  # 左侧白边，不能有任何占用的区域
                "right_white": 10,
                "left": 275,
                "up": 90,
                "down": 90,
                "right": 150,
                "label_width": 80,
                "mile_label_width": 40,
                "ruler_label_width": 80,
            },
            "type_regex": [
                ('高速', r'G\d+', True),
                ('动车组', r'D\d+', True),
                ('城际', r'C\d+', True),
                ('直达特快', r'Z\d+', True),
                ('特快', r'T\d+', True),
                ('快速', r'K\d+', True),
                ('普快', r'[1-5]\d{3}$', True),  # 非复用
                ('普快', r'[1-5]\d{3}\D', True),  # 复用
                ('普客', r'6\d{3}$', True),
                ('普客', r'6\d{3}\D', True),
                ('普客', r'7[1-5]\d{2}$', True),
                ('普客', r'7[1-5]\d{2}\D', True),
                ('通勤', r'7\d{3}$', True),
                ('通勤', r'7\d{3}\D', True),
                ('通勤', r'8\d{3}$', True),
                ('通勤', r'8\d{3}\D', True),
                ('旅游', r'Y\d+', True),
                ('路用', r'57\d+', True),
                ('特快行包', r'X1\d{2}', True),
                ('动检', r'DJ\d+', True),
                ('客车底', r'0[GDCZTKY]\d+', True),
                ('临客', r'L\d+', True),  # 主要解决类型直接定为“临客”的车次。
                ('客车底', r'0\d{4}', True),
                ('行包', r'X\d{3}\D', False),
                ('行包', r'X\d{3}$', False),
                ('班列', r'X\d{4}', False),
                ('直达', r'1\d{4}', False),
                ('直货', r'2\d{4}', False),
                ('区段', r'3\d{4}', False),
                ('摘挂', r'4[0-4]\d{3}', False),
                ('小运转', r'4[5-9]\d{3}', False),
                ('单机', r'5[0-2]\d{3}', False),
                ('补机', r'5[3-4]\d{3}', False),
                ('试运转', r'55\d{3}', False),
            ]  # 类型对应正则表达式，list<tuple>有序三元列表。数据结构为类型，正则，是否客车。优先级递减。
        }
        default_config.update(self._sysConfig)
        self._sysConfig = default_config

    def saveSysConfig(self):
        """
        1.4版本新增函数。保证进入本函数时系统设置是有效的。
        """
        with open(config_file, 'w', encoding='utf-8', errors='ignore') as fp:
            json.dump(self._sysConfig, fp, ensure_ascii=False)

    def initGraphConfig(self):
        """
        1.4新增函数。在不考虑读文件的前提下，增加单个运行图设置应该比系统运行图设置多的东西。
        precondition: sysConfig已经读取并校验完毕。
        """
        self._config = {
            "ordinate": None,
            "not_show_types": [],
        }
        self._config.update(self._sysConfig)

    def checkGraphConfig(self):
        """
        1.4新增函数，由loadGraph调用（初始化时可以用更加暴力的方法）。
        检查并更新graphConfig。precondition: self._config存在且是dict；sysConfig已经初始化并检查完毕。
        """
        self._config.setdefault("ordinate", None)
        self._config.setdefault("not_show_types", [])
        for key, value in self._sysConfig.items():
            self._config.setdefault(key, value)

    def resetGraphConfigFromConfigWidget(self):
        """
        1.4版本新增函数。提供给configWidget调用的重置接口。
        precondition: self._config和self._sysConfig都是合法的。
        """
        # for key,value in self._sysConfig:
        #     if 'color' not in key:
        #         self._config[key]=value
        self._config.update(self._sysConfig)

    def setFullCheciMap(self):
        """
        设置全车次查找表。
        """
        for train in self.trains():
            self.fullCheciMap[train.fullCheci()] = train

    def setSingleCheciMap(self):
        for train in self.trains():
            for cc in (train.downCheci(), train.upCheci()):
                if cc:
                    self.singleCheciMap.setdefault(cc, []).append(train)

    def addSingleCheciMap(self, train):
        for cc in (train.downCheci(), train.upCheci()):
            if cc:
                self.singleCheciMap.setdefault(cc, []).append(train)

    def delSingleCheciMap(self, train):
        for cc in (train.downCheci(), train.upCheci()):
            if cc:
                lst = self.singleCheciMap.get(cc)
                if len(lst) > 1:
                    lst.remove(train)
                else:
                    del self.singleCheciMap[cc]

    def changeTrainCheci(self, train: Train, full: str, down: str, up: str):
        """
        2019.02.24新增加，更新车次，并更新查找表。
        """
        f, d, u = train.checi
        if full != f and self.fullCheciMap.get(f, None) is not None:
            del self.fullCheciMap[f]
            self.fullCheciMap[full] = train
        reset = False
        if d != down or u != up:
            self.delSingleCheciMap(train)
            reset = True
        train.setCheci(full, down, up)
        if reset:
            self.addSingleCheciMap(train)

    def refreshData(self):
        """
        2019.02.24新增。系统刷新时强制刷新所有映射表数据，保证正常工作。
        """
        self.setFullCheciMap()
        self.setSingleCheciMap()
        self.line.setNameMap()
        self.line.setFieldMap()

    def loadGraph(self, filename: str):
        """
        暂定直接打开json文件读
        """
        self.filename = filename
        fp = open(filename, encoding='utf8', errors='ignore')
        try:
            info = json.load(fp)
        except json.JSONDecodeError:
            fp.close()
            self.loadTrcGraph(filename)
            return

        self.line.loadLine(info["line"])
        self._circuits = []
        for dict_train in info["trains"]:
            newtrain = Train(self, origin=dict_train)
            self._trains.append(newtrain)

        # 2020.07.01：调整顺序。必须先初始化好map，再读交路。
        self.setFullCheciMap()
        self.setSingleCheciMap()

        for c in info.get('circuits', []):
            self._circuits.append(Circuit(self, origin=c))

        self._config = info.get("config", {})
        if not isinstance(self._config, dict):
            self._config = {}
        if self._config.get('ordinate') is not None:
            self._config["ordinate"] = self.line.rulerByName(self._config["ordinate"])
        self.checkGraphConfig()

        self._markdown = info.get("markdown", '')
        try:
            self._version = info['version']
        except KeyError:
            pass

        fp.close()

    def version(self) -> str:
        return self._version

    def setVersion(self, v: str):
        self._version = v

    def sysVersion(self) -> str:
        if self._sysVersion:
            return self._sysVersion
        return self._version

    def setSysVersion(self, v: str):
        self._sysVersion = v

    def addTrain(self, train: Train):
        train.graph = self
        self._trains.append(train)
        self.fullCheciMap[train.fullCheci()] = train
        self.addSingleCheciMap(train)

    def delTrain(self, train: Train):
        # import traceback
        try:
            if train.carriageCircuit() is not None:
                train.carriageCircuit().changeTrainToVirtual(train)
            self._trains.remove(train)
            del self.fullCheciMap[train.fullCheci()]
            self.delSingleCheciMap(train)
        except Exception as e:
            print("del train: No such train!", train.fullCheci())
            # traceback.print_exc()

    def trains(self):
        for train in self._trains:
            yield train

    def show(self):
        self.line.show()
        # print(self.circuits)
        for train in self._trains:
            train.show()

    def setLine(self, line: Line):
        self.line = line

    def lineLength(self)->float:
        return self.line.lineLength()

    def counterLength(self)->float:
        """
        [对里程]意义下的线路长度，或者说是上行线的长度。仅考虑最后一个站。
        如果最后一个站的对里程数据不存在，则使用正里程长度数据。
        """
        return self.line.counterLength()

    def stations(self, reverse=False):
        if not reverse:
            for station in self.line.stations:
                yield station["zhanming"]
        else:
            for station in reversed(self.line.stations):
                yield station["zhanming"]

    def businessStationNames(self, passenger_only: bool, freight_only: bool):
        """
        2.1.1版本新增加。对办客和办货做筛选，由intervalCount过程调用。
        由于stations()函数调用范围太广，怕引起其他问题，故新增本函数。
        """
        for st in self.line.stations:
            if ((not passenger_only) or st.get("passenger", True)) and \
                    ((not freight_only) or st.get("freight", True)):
                yield st['zhanming']

    def verifyStationBusiness(self, st: dict, passenger_only: bool, freight_only: bool) -> bool:
        """
        2019.07.12新增。
        检查所给的车站dict是否符合办客和办货的要求。
        """
        if ((not passenger_only) or st.get("passenger", True)) and \
                ((not freight_only) or st.get("freight", True)):
            return True
        return False

    def stationDicts(self, reverse=False):
        if not reverse:
            for station in self.line.stations:
                yield station
        else:
            for station in reversed(self.line.stations):
                yield station

    def save(self, filename: str):
        """
        保存运行图文件
        """
        graph = {
            "line": self.line.outInfo(),
            "trains": [],
            "circuits": [],
            "config": self._config,
            "version": self._version,
        }
        for c in self._circuits:
            graph['circuits'].append(c.outInfo())
        try:
            graph["markdown"] = self._markdown
        except AttributeError:
            self._markdown = ''
            graph["markdown"] = ''
        if graph["config"]["ordinate"] is not None:
            if isinstance(graph["config"]["ordinate"], str):
                pass
            else:
                graph["config"]["ordinate"] = graph["config"]["ordinate"].name()

        for train in self._trains:
            graph["trains"].append(train.outInfo())

        with open(filename, 'w', encoding='utf8', errors='ignore') as fp:
            # print(graph["line"]["rulers"])
            json.dump(graph, fp, ensure_ascii=False)

    def UIConfigData(self):
        return self._config

    def sysConfigData(self):
        """
        1.4版本新增函数，专用返回系统默认设置。
        """
        return self._sysConfig

    def line_station_mileages(self):
        """
        生成器，返回 dict["zhanming"],dict["licheng"]
        """
        for dict in self.line.stations:
            yield dict["zhanming"], dict["licheng"]

    def line_station_mile_levels(self):
        for dict in self.line.stations:
            yield dict["zhanming"], dict["licheng"], dict["dengji"]

    def addEmptyRuler(self, name: str, different: bool = False):
        ruler = Ruler(name=name, different=different, line=self.line)
        self.line.rulers.append(ruler)
        return ruler

    def addRuler(self, ruler: Ruler):
        self.line.rulers.append(ruler)

    def delRuler(self, ruler: Ruler):
        """
        返回被删除的标尺是否是排图标尺
        """
        try:
            self.line.rulers.remove(ruler)
        except ValueError:
            pass
        if self.ordinateRuler() is ruler:
            self.setOrdinateRuler(None)
            return True
        return False

    def setOrdinateRuler(self, ruler: Ruler):
        self._config["ordinate"] = ruler

    def ordinateRuler(self):
        """
        2019.05.09：有时候莫名其妙会是一个str对象，就暴力解决问题。
        """
        ruler = self._config.setdefault("ordinate", None)
        if isinstance(ruler, str):
            print("Graph::ordinateRuler: ruler is str", ruler)
            return self.line.rulerByName(ruler)
        return ruler

    def setStationYValue(self, name, y):
        """
        设置某个站的纵坐标值。2019.02.02移植到line类并删除线性算法。
        """
        self.line.setStationYValue(name, y)

    def stationYValue(self, name: str):
        st_dict = self.stationByDict(name)
        if st_dict is not None:
            return st_dict.get('y_value', -1)
        else:
            return -1

    def stationMileYValues(self):
        for dct in self.line.stations:
            yield dct['zhanming'], dct['licheng'], dct.get('y_value', None)

    def trainFromCheci(self, checi: str, full_only=False) -> Train:
        """
        根据车次查找Train对象。如果full_only，则仅根据全车次查找；否则返回单车次匹配的第一个结果。
        若不存在，返回None。
        2019.02.03删除线性算法。
        注意：调用之前必须保证fullCheciMap已经初始化！
        """
        t = self.fullCheciMap.get(checi, None)
        if t is not None:
            return t
        if not full_only:
            selected = self.singleCheciMap.get(checi, None)
            if selected is not None:
                return selected[0]
        return None
        # t = self.trainFromCheciLinear(checi,full_only)
        # if t is not None:
        #     print("Graph::trainFromCheci error! Existed train not found: ",t)
        # return t

    def trainFromCheciLinear(self, checi: str, full_only: bool) -> Train:
        """
        仅调试用，线性查找车次。
        """
        # print("Graph::trainFromCheciLinear",checi)
        for train in self.trains():
            if train.fullCheci() == checi:
                return train
        if not full_only:
            for train in self.trains():
                if train.downCheci() == checi or train.upCheci() == checi:
                    return train
        return None

    def multiSearch(self, checi: str):
        """
        非严格搜索。线性算法。
        """
        selected = []
        for train in self.trains():
            if checi in train.fullCheci() or checi in train.downCheci() or checi in train.upCheci():
                selected.append(train)
        return selected

    def lineName(self):
        return self.line.name

    def setLineName(self, name: str):
        self.line.name = name

    def firstStation(self):
        if self.line.stations:
            return self.line.stations[0]["zhanming"]
        else:
            return None

    def lastStation(self):
        try:
            return self.line.stations[-1]["zhanming"]
        except IndexError:
            return None

    def stationInLine(self, name: str, strict=False):
        return self.line.stationInLine(name, strict)

    def setStationIsShow(self, name: str, show: bool):
        """
        不支持域解析符。2019.02.02删除线性算法。
        2019.07.30：似无调用
        """
        st = self.line.stationDictByName(name, True)
        if st is None:
            raise Exception("setStationIsShow: no such station")
        else:
            st['show'] = show

    def stationIsShow(self, name: str):
        dct = self.line.stationDictByName(name)
        if dct is None:
            raise Exception("No such station")
        return dct.setdefault('show', True)

    def graphFileName(self):
        return self.filename

    def setGraphFileName(self, filename: str):
        self.filename = filename

    def isEmpty(self):
        if self.line.stations:
            return False
        else:
            return True

    def stationCount(self):
        return len(self.line.stations)

    def clearLineStationInfo(self):
        self.line.clear()

    def stationExisted(self, name: str)->bool:
        return self.line.stationExisted(name)

    def addStationDict(self, info: dict):
        self.line.stations.append(info)

    def adjustLichengTo0(self):
        if self.isEmpty():
            return
        start_mile = self.line.stations[0]["licheng"]
        for st in self.line.stations:
            st["licheng"] = st["licheng"] - start_mile

    def trainCount(self):
        return len(self._trains)

    def rulers(self):
        for ruler in self.line.rulers:
            yield ruler

    def gapBetween(self, st1: str, st2: str)->float:
        """
        计算两个站间距离.
        2020.01.23新增：如果是上行方向，则尝试使用对里程。
        对里程按照点对点原则使用，只考虑两端点的对里程数据，不考虑中间的。
        """
        return self.line.gapBetween(st1,st2)

    def lineSplited(self):
        """
        返回本线是否存在上下行分设站的情况
        """
        if self.line.isSplited():
            return True
        return False

    def rulerNameExisted(self, name, ignore: Ruler = None):
        for r in self.line.rulers:
            if r is not ignore and r.name() == name:
                return True
        return False

    def validRulerName(self)->str:
        s = "新标尺0"
        i = 0
        while self.rulerNameExisted(s):
            i+=1
            s = f"新标尺{i}"
        return s

    def circuitNameExisted(self, name, ignore: Circuit = None):
        for c in self.circuits():
            if c is not ignore and c.name() == name:
                return True
        return False

    def validNewCircuitName(self)->str:
        """
        以新建x的格式返回一个有效的新建标尺名称。
        """
        s = "新建0"
        i = 0
        while self.circuitNameExisted(s):
            i+=1
            s = f"新建{i}"
        return s

    def isNewRuler(self, ruler: Ruler):
        for r in self.line.rulers:
            if ruler is r:
                return False
        return True

    def stationDirection(self, name: str):
        return self.line.stationViaDirection(name)

    def lineStationBusiness(self, name: str, passenger: int, default=False) -> bool:
        """
        2.0.2新增，返回车站是否办理业务。passenger是Train中规定的枚举常量，标志是否办客。
        如果找不到，返回default。
        """
        dct = self.line.stationDictByName(name)
        if dct is None:
            # print("graph::lineStationBusiness: no such station! return",default,name)
            return default

        if passenger == Train.PassengerTrue:
            # print("graph::lineStationBusiness passengerTrue",dct.get('passenger',"无数据"))
            return dct.get('passenger', True)
        else:
            # print("graph::lineStationBusiness passengerFalse",dct.get("freight","无数据"))
            return dct.get("freight", True)

    def formerBothStation(self, name: str):
        """
        寻找本站往前的第一个【下行方向通过】的站。
        TODO 2019.02.02 保留线性算法。下一个函数同。
        """
        former_dict = None
        for st in self.line.stations:
            if st["zhanming"] == name:
                return former_dict

            if st["direction"] == 0x3:
                former_dict = st
        raise Exception("No former station")

    def latterBothStation(self, name: str):
        start = False
        for st in self.line.stations:
            if st["zhanming"] == name:
                start = True
            if start and st["direction"] == 0x3:
                return st
        raise Exception("No latter station")

    def stationLevel(self, name: str):
        """
        返回车站等级。若不存在，返回None；若没有这个字段，设为并返回4. 不支持域解析符
        """
        st = self.line.stationDictByName(name, strict=True)
        if st is None:
            return None
        return st.setdefault('dengji', 4)

    def setNotShowTypes(self, not_show):
        self.UIConfigData()["not_show_types"] = not_show
        for train in self.trains():
            if train.type in not_show:
                train.setIsShow(False, affect_item=False)
            else:
                train.setIsShow(True, affect_item=False)

    def setDirShow(self, down, show, itemWise=False):
        """
        :param itemWise  是否精确到每一段运行线。True-按照每一段运行线的上下行组织，可能导致运行线的部分显示；False-仅按照入图上下行。
        """
        for train in self.trains():
            if train.type not in self.UIConfigData()['not_show_types']:
                if itemWise:
                    # 如果要显示，且没有铺画过，必须强制设成要显示，以免漏掉
                    if show and not train._itemInfo:
                        train.setIsShow(True, affect_item=False)
                    elif show:
                        # 显示所有包含指定方向的
                        if not train.isShow():
                            for info in train.itemInfo():
                                if info['down'] == down:
                                    train.setIsShow(True, affect_item=False)
                                    break
                    else:  # not show
                        if train.isShow():
                            foundNonThis = False
                            for info in train.itemInfo():
                                if info['down'] != down:
                                    foundNonThis = True
                                    break
                            if not foundNonThis:
                                # 不显示【所有运行线全为本方向】的车次
                                train.setIsShow(False, affect_item=False)

                else:
                    if train.firstDown() == down:
                        train.setIsShow(show, affect_item=False)

    def trainExisted(self, train: Train, ignore: Train = None):
        """
        比较Train对象，线性算法
        """
        for t in self._trains:
            if train is t and t is not ignore:
                return True
        return False

    def checiExisted(self, checi: str, ignore: Train = None):
        """
        比较全车次。2019.02.03替换掉线性算法。
        """
        # for t in self._trains:
        #     if t is not ignore and t.fullCheci() == checi:
        #         return True
        # return False
        t = self.fullCheciMap.get(checi, None)
        if t is not None and (ignore is None or t is not ignore):
            return True
        return False

    def rulerCount(self):
        return len(self.line.rulers)

    def checiType(self, checi: str) -> str:
        """
        2.0.2新增。根据系统设置的判断规则，返回车次对应的类型。如果不符合任何一个，返回 其他。
        """
        for nm, rg, _ in self.UIConfigData()['type_regex']:
            try:
                rgx = re.compile(rg)
            except:
                print("Invalid Regex! ", rg)
                continue
            if re.match(rgx, checi):
                return nm
        return '其他'

    def checiTypePassenger(self, checi: str) -> (str, int):
        """
        根据车次返回类型以及是否是客车。是否是客车按照Train中定义的常量。
        如果不符合任何一个，返回 其他, PassengerAuto。
        """
        for nm, rg, ps in self.UIConfigData()['type_regex']:
            if re.match(rg, checi):
                if ps:
                    return nm, Train.PassengerTrue
                else:
                    return nm, Train.PassengerFalse
        return '其他', Train.PassengerAuto

    def typePassenger(self, tp: str, default=Train.PassengerAuto) -> int:
        """
        根据类型返回是否为客车。返回是Train中的PassengerTrue或PassengerFalse，如果找不到返回默认。
        """
        for name, _, ps in self.UIConfigData()['type_regex']:
            if name == tp:
                if ps:
                    return Train.PassengerTrue
                else:
                    return Train.PassengerFalse
        return default

    def stationTimeTable(self, name: str):
        """
        返回车站的图定时刻表
        list<dict>
        dict{
            "station_name":str,
            "ddsj":datetime,
            "cfsj":datetime,
            "down":bool,
            "train":Train,
            "track": !str,  //2020.01.24新增股道
        }
        """
        timeTable = []
        for train in self.trains():
            st_dict = train.stationDict(name)
            if st_dict is None:
                continue
            else:
                node = {
                    "ddsj": st_dict["ddsj"],
                    "cfsj": st_dict["cfsj"],
                    "station_name": st_dict["zhanming"],
                    "down": train.stationDown(st_dict['zhanming'], self),
                    "note": st_dict.get("note", ''),
                    "train": train,
                    "track":st_dict.get("track",None),
                }
                timeTable.append(node)

        # 排序
        for i in range(len(timeTable) - 1):
            t = i
            for j in range(i + 1, len(timeTable)):
                if timeTable[j]["ddsj"] < timeTable[t]["ddsj"]:
                    t = j
            temp = timeTable[t]
            timeTable[t] = timeTable[i]
            timeTable[i] = temp
        return timeTable

    def reverse(self):
        """
        反排运行图。
        2020年1月24日将线路部分逻辑封装到line里。
        """
        self.line.reverse()

        # 列车上下行调整、上下行车次交换
        for train in self._trains:
            # 上下行交换
            train.reverseAllItemDown()

            # 车次交换
            train.setCheci(train.fullCheci(), train.upCheci(), train.downCheci())

    def downTrainCount(self):
        count = 0
        for train in self.trains():
            if train.firstDown() is True:
                count += 1
        return count

    def upTrainCount(self):
        count = 0
        for train in self.trains():
            if train.firstDown() is False:
                count += 1
        return count

    def loadTrcGraph(self, filename):
        """
        阅读旧版trc格式的运行图
        """
        fp = open(filename, encoding='utf-8', errors='ignore')
        self.line.forbid.setDifferent(False)
        self.line.forbid.setShow(True, True)
        inTrainArea = False
        now_list = []
        last_name = None
        circuit_dict = {}
        for i, line in enumerate(fp):
            line = line.strip()
            if not line:
                continue
            if not inTrainArea and line == "===Train===":
                inTrainArea = True

            if line[0] == '-':
                break

            # 处理线路信息部分
            if not inTrainArea:
                if line == "***Circuit***":
                    continue
                elif i == 1:
                    self.setLineName(line)
                else:
                    # 线路信息部分
                    try:
                        splited = line.split(',')
                        st_name = splited[0]
                        self.line.addStation_by_info(splited[0], int(splited[1]), int(splited[2]))
                        if last_name is not None:
                            try:
                                start_str, end_str = splited[9].split('-', 1)
                                begin = datetime.strptime(start_str, '%H:%M')
                                end = datetime.strptime(end_str, '%H:%M')
                                self.line.forbid.addForbid(last_name, st_name, begin, end)
                            except Exception as e:
                                pass
                        last_name = st_name
                    except:
                        pass

            # 处理列车信息部分
            else:
                # 这部分从trc_check_new中复制过来
                if line != '===Train===':
                    now_list.append(line)
                else:
                    self._decodeTrcTrain(now_list, circuit_dict)
                    now_list = []
        self._decodeTrcTrain(now_list, circuit_dict)
        self._decodeTrcCircuit(circuit_dict)
        self.setGraphFileName('')

    def _decodeTrcTrain(self, now_list: list, circuit_dict: dict):
        """
        阅读trc中单个车次的信息，不含===Train===标志头。
        circuit_dict: 抽取车次中含有的交路信息。数据结构为
        Dict<List<Tuple<int,Train>>>
        eg.
        {
            "CRH380-2081":[
                (0, Train<G1>),
                (1, Train<G4>),
                ...
            ],
            ...
        }
        交路数据保证只有一个下划线“_”，且split之后是一个整数。
        """
        train = Train(self)
        for i, line in enumerate(now_list):
            if i == 0:
                splited = line.split(',')
                train.setCheci(splited[1], splited[2], splited[3])
                if len(splited) >= 5:
                    circuit_str = splited[4]
                    try:
                        num = int(circuit_str.split('_')[-1])
                        name = circuit_str.split('_')[0]
                    except ValueError:
                        if circuit_str not in ('NA', ''):
                            print("Graph::decodeTrcTrain: Unexpected circuit info:", circuit_str)
                    else:
                        circuit_dict.setdefault(name, []).append((num, train))


            elif i == 1:
                train.setStartEnd(sfz=line)
            elif i == 2:
                train.setStartEnd(zdz=line)
            else:
                splited = line.split(',')
                train.addStation(splited[0], splited[1], splited[2])
        train.autoTrainType()
        if train.timetable:
            self.addTrain(train)

    def _decodeTrcCircuit(self, circuit_dict: dict):
        """
        解析从前面收集到的交路数据，生成交路对象。
        """
        for name, lst in circuit_dict.items():
            lst.sort()
            circuit = Circuit(self, name)
            for _, train in lst:
                circuit.addTrain(train)
            self.addCircuit(circuit)

    def jointGraph(self, graph, former: bool, reverse: bool, line_only: bool):
        """
        拼接两运行图。
        :param graph: 另一运行图
        :param former: 另一运行图是否在本运行图前侧链接
        :param reverse: 另一运行图是否转置
        """
        if reverse:
            graph.reverse()

        if not line_only:
            # 车次连接
            for train_append in graph.trains():
                if self.checiExisted(train_append.fullCheci()):
                    # 本线有相同车次
                    train_main: Train = self.trainFromCheci(train_append.fullCheci())
                    train_main.delNonLocal(self)
                    train_append.delNonLocal(graph)
                    # 方向以本线为准
                    # down表示列车在两线路连接点附近的上下行情况。
                    if former:
                        down = train_main.firstDown()
                    else:
                        down = train_main.lastDown()
                    if down is None:
                        # 如果本线无法判断，例如Z90终到石家庄在京石段只有一个站，则用另一条线的。
                        if former:
                            down = train_append.lastDown()
                        else:
                            down = train_append.firstDown()
                    if down is None:
                        # 如果都无法判断，直接判断为下行车次
                        print("cannot judge down. use default.", train_main.fullCheci())
                        down = True

                    train_former = not (down ^ former)
                    train_main.jointTrain(train_append, train_former, graph)  # 当前站点已经拼接好

                else:
                    self.addTrain(train_append)

        # 线路连接
        self.line.jointLine(graph.line,former,reverse)

    def resetAllItems(self):
        for train in self.trains():
            train.setItem(None)

    def stationMile(self, name: str):
        """
        返回车站的里程数据，若不存在返回-1.不支持域解析符。2019.02.03删除线性算法。
        2019.02.23改为：支持域解析符。
        """
        st = self.line.stationDictByName(name)
        if st is None:
            return -1
        return st["licheng"]

    def adjacentStation(self, name: str, ignore: list):
        index = self.stationIndex(name)
        if index > 0:
            if self.line.stations[index - 1]['zhanming'] not in ignore:
                # 2019.02.23修改，条件少了zhanming，not in的判断相当于没用
                return self.line.stations[index - 1]["zhanming"]

        if index < len(self.line.stations) - 1:
            if self.line.stations[index + 1]['zhanming'] not in ignore:
                return self.line.stations[index + 1]["zhanming"]
        print("no adj")
        return None

    def stationIndex(self, name: str):
        """
        2019.07.12新增常量级别算法。理论上应保证站名存在。
        """
        if self.line.numberMap is None:
            return self.stationIndex_bf(name)
        else:
            try:
                return self.line.numberMap[self.nameMapToLine(name)]
            except KeyError:
                print("Graph::stationIndex: Unexpected station name:", name)
                return self.stationIndex_bf(name)

    def stationIndex_bf(self, name: str):
        """
        原来的暴力方法查找序号。分离此函数是为了尝试统计有多少次使用暴力方法。
        """
        for i, st in enumerate(self.line.stations):
            if stationEqual(st["zhanming"], name):
                return i
        raise StationNotInLineException(name)

    def stationByDict(self, name: str, strict=False)->LineStation:
        """
        根据站名返回dict对象，函数名写错了。支持域解析符。
        2019.02.02删除线性算法。
        """
        return self.line.stationDictByName(name, strict)

    def passedStationCount(self, st1: str, st2: str, down: bool) -> int:
        """
        检查以st1为发站，st2为到站，方向为down的区间内有多少个站。2.0新增。
        """
        s1 = self.stationIndex(st1)
        s2 = self.stationIndex(st2)
        dir_ = 0b1 if down else 0b10
        cnt = 0
        t1 = min((s1, s2))
        t2 = max((s1, s2))
        # print("t1 t2",t1,t2)
        for i in range(t1 + 1, t2):
            dct = self.line.stationDictByIndex(i)
            if dir_ & dct.get('direction', 0b11):
                cnt += 1
        # print("passedStationCount",st1,st2,down,cnt)
        return cnt

    def resetStationName(self, old, new, auto_field=False):
        old_dict = self.stationByDict(old)
        if old_dict is not None:
            old_dict["zhanming"] = new if not auto_field else new.split('::')[0]

        # 更新标尺中站名
        for ruler in self.line.rulers:
            ruler.changeStationName(old, new)
        for train in self.trains():
            if train.isSfz(old):
                train.sfz = new
            if train.isZdz(old):
                train.zdz = new
            if train._localFirst == old:
                train._localFirst = new
            elif train._localLast == old:
                train._localLast = new
            st_dict = train.stationDict(old)
            if st_dict is not None:
                st_dict["zhanming"] = new
        # 更新天窗中站名
        self.line.forbid.changeStationName(old, new)
        self.line.forbid2.changeStationName(old, new)
        self.line.changeStationNameUpdateMap(old, new)

    def addTrainByGraph(self, graph, cover=False):
        """
        添加车次，返回数量
        """
        num = 0
        for train in graph.trains():
            if train.localCount(self) >= 2:
                if not self.checiExisted(train.fullCheci()):
                    num += 1
                    self.addTrain(train)
                elif cover:
                    num += 1
                    t = self.trainFromCheci(train.fullCheci())
                    # 临时处理：移交交路数据
                    circuit = t.carriageCircuit()
                    if circuit is not None:
                        circuit.replaceTrain(t, train)
                        train.setCarriageCircuit(circuit)
                    self.delTrain(t)
                    self.addTrain(train)
        self.checkCircuits()
        return num

    def preAddTrainByGraph(self, graph, all:bool=False):
        """
        2019.07.19新增。预导入所有与本线有关涉的车次和交路。
        此函数只能在临时对象中调用。自身的车次表、交路表应当是空的。
        注意，此操作后的circuits是不安全的，执行train()可能会引发TrianNotFoundException。
        """
        tm1 = time.perf_counter()
        for train in graph.trains():
            # if train.localCount(self) >= 2:
            if all or train.isLocalTrain(self):
                circuit = train.carriageCircuit()
                if circuit is not None:
                    if circuit not in self._circuits:
                        circuit.setGraph(self)
                        self.addCircuit(circuit)
                self.addTrain(train)
        tm2 = time.perf_counter()
        print("预导入线路历时", tm2 - tm1)

    def checkCircuits(self):
        """
        2020.02.02新增。
        检查所有交路信息。如果找不到对应的车次，则设置为虚拟。
        """
        for circuit in self.circuits():
            circuit.identifyTrain(full_only=True)

    def setMarkdown(self, mark: str):
        self._markdown = mark

    def markdown(self):
        try:
            return self._markdown
        except AttributeError:
            self._markdown = ""
            return ''

    def save_excel(self, filename: str):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = f'{self.firstStation()}-{self.lastStation()}间列车时刻表'

        # 写入左侧表头
        ws['A3'] = '始发站'
        ws.merge_cells('A3:A4')
        ws['A5'] = '终到站'
        ws.merge_cells('A5:A6')
        ws['A7'] = '列车种类'
        ws.merge_cells('A7:A8')
        ws['A9'] = '车次'
        ws['A10'] = '车站'
        for row in range(3, 11):
            ws.row_dimensions[row].font = Font(name='SimSum', size=9)
            ws.row_dimensions[row].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 9.7

        start = 11  # 从第11行开始表格
        # 写入车站
        station_row_dict = {}
        cur = 11
        for station in self.stations():
            ws.cell(row=cur, column=1, value=station)
            ws.merge_cells(start_row=cur, end_row=cur + 1, start_column=1, end_column=1)
            station_row_dict[station] = cur
            ws.row_dimensions[cur].height = 9.7
            ws.row_dimensions[cur + 1].height = 9.7
            ws.row_dimensions[cur].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[cur + 1].alignment = Alignment(horizontal='center', vertical='center')
            cur += 2
        ws.column_dimensions['A'].width = 12

        # 写入车次，先下行
        last_merge_sfz, last_merge_zdz, last_merge_type = 1, 1, 1
        col = 2
        last_train = None
        for train in self.trains():
            for dct in train.itemInfo():
                if not dct['down']:
                    continue
                if last_train and train.sfz == last_train.sfz:
                    try:
                        ws.unmerge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col)
                else:
                    ws.merge_cells(start_row=3, end_row=4, start_column=col, end_column=col)
                    last_merge_sfz = col
                ws.cell(row=3, column=last_merge_sfz, value=train.sfz)  # 必须访问最左边的才行

                if last_train and train.zdz == last_train.zdz:
                    try:
                        ws.unmerge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col)
                else:
                    ws.merge_cells(start_row=5, end_row=6, start_column=col, end_column=col)
                    last_merge_zdz = col
                c = ws.cell(row=5, column=last_merge_zdz, value=train.zdz)
                col_str = c.column_letter
                ws.column_dimensions[col_str].width = 6  # 设置列宽为5

                if last_train and train.type == last_train.type:
                    try:
                        ws.unmerge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col)
                else:
                    ws.merge_cells(start_row=7, end_row=8, start_column=col, end_column=col)
                    last_merge_type = col
                ws.cell(row=7, column=last_merge_type, value=train.type)

                checi = train.fullCheci()
                if '/' in checi:
                    ws.cell(row=9, column=col, value=checi.split('/')[0])
                    ws.cell(row=10, column=col, value='/' + checi.split('/', maxsplit=1)[1])
                else:
                    ws.cell(row=9, column=col, value=checi)
                    ws.merge_cells(start_row=9, end_row=10, start_column=col, end_column=col)

                last_dict = None

                # 时刻表循环
                for st_dict in train.timetable:
                    for i, s in station_row_dict.items():
                        if stationEqual(i, st_dict['zhanming']):
                            row = s
                            break
                    else:
                        continue

                    if train.isSfz(st_dict['zhanming']):
                        ws.cell(row=row, column=col, value='')
                        ws.cell(row=row + 1, column=col, value=self.outTime(st_dict['cfsj'], True))

                    elif train.isZdz(st_dict["zhanming"]):
                        ws.cell(row=row, column=col, value=self.outTime(st_dict['ddsj'], True))
                        ws.cell(row=row + 1, column=col, value='    --')

                    elif train.stationStopped(st_dict):
                        # 本站停车，无条件写入完整到达时刻和不完整出发时刻
                        ddsj_str = f'{st_dict["ddsj"].hour:2d}:{st_dict["ddsj"].minute:02d}'
                        sec = st_dict['ddsj'].second
                        if sec:
                            ddsj_str += f"{sec:02d}"
                        else:
                            ddsj_str += '  '
                        ws.cell(row=row, column=col, value=ddsj_str)
                        if st_dict['ddsj'].hour == st_dict['cfsj'].hour:
                            cfsj_str = '   '
                        else:
                            cfsj_str = f"{st_dict['cfsj'].hour:2d}:"
                        cfsj_str += f'{st_dict["cfsj"].minute:02d}'
                        sec = st_dict['cfsj'].second
                        if sec:
                            ddsj_str += f"{sec:02d}"
                        else:
                            ddsj_str += '  '
                        ws.cell(row=row + 1, column=col, value=cfsj_str)

                    else:
                        give_hour = False
                        if not last_dict:
                            give_hour = True
                        elif last_dict['cfsj'].hour != st_dict['ddsj'].hour:
                            give_hour = True
                        ws.cell(row=row, column=col, value='   ...')
                        tgsj_str = f'{st_dict["ddsj"].hour:2d}:' if give_hour else '   '
                        tgsj_str += f'{st_dict["ddsj"].minute:02d}'
                        sec = st_dict['ddsj'].second
                        if sec:
                            tgsj_str += f"{sec:02d}"
                        else:
                            tgsj_str += '  '
                        ws.cell(row=row + 1, column=col, value=tgsj_str)
                    last_dict = st_dict
                col += 1
                last_train = train

        # 上行
        for train in self.trains():
            for dct in train.itemInfo():
                if dct['down']:
                    continue
                if last_train and train.sfz == last_train.sfz:
                    try:
                        ws.unmerge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col)
                else:
                    ws.merge_cells(start_row=3, end_row=4, start_column=col, end_column=col)
                    last_merge_sfz = col
                c = ws.cell(row=3, column=last_merge_sfz, value=train.sfz)
                col_str = c.column_letter
                ws.column_dimensions[col_str].width = 6  # 设置列宽为5

                if last_train and train.zdz == last_train.zdz:
                    try:
                        ws.unmerge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col)
                else:
                    ws.merge_cells(start_row=5, end_row=6, start_column=col, end_column=col)
                    last_merge_zdz = col
                ws.cell(row=5, column=last_merge_zdz, value=train.zdz)

                if last_train and train.type == last_train.type:
                    try:
                        ws.unmerge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col - 1)
                    except:
                        pass
                    ws.merge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col)
                else:
                    ws.merge_cells(start_row=7, end_row=8, start_column=col, end_column=col)
                    last_merge_type = col
                ws.cell(row=7, column=last_merge_type, value=train.type)

                checi = train.fullCheci()
                if '/' in checi:
                    ws.cell(row=9, column=col, value=checi.split('/')[0])
                    ws.cell(row=10, column=col, value='/' + checi.split('/', maxsplit=1)[1])
                else:
                    ws.cell(row=9, column=col, value=checi)
                    ws.merge_cells(start_row=9, end_row=10, start_column=col, end_column=col)

                last_dict = None
                # 时刻表循环
                for st_dict in train.timetable:
                    for i, s in station_row_dict.items():
                        if stationEqual(i, st_dict['zhanming']):
                            row = s
                            break
                    else:
                        continue

                    if train.isSfz(st_dict['zhanming']):
                        ws.cell(row=row + 1, column=col, value='')
                        ws.cell(row=row, column=col, value=self.outTime(st_dict['cfsj'], True))

                    elif train.isZdz(st_dict["zhanming"]):
                        ws.cell(row=row + 1, column=col, value=self.outTime(st_dict['ddsj'], True))
                        ws.cell(row=row, column=col, value='    --')

                    elif train.stationStopped(st_dict):
                        # 本站停车，无条件写入完整到达时刻和不完整出发时刻
                        ddsj_str = f'{st_dict["ddsj"].hour:2d}:{st_dict["ddsj"].minute:02d}'
                        sec = st_dict['ddsj'].second
                        if sec:
                            ddsj_str += f"{sec:02d}"
                        else:
                            ddsj_str += '  '
                        ws.cell(row=row + 1, column=col, value=ddsj_str)
                        if st_dict['ddsj'].hour == st_dict['cfsj'].hour:
                            cfsj_str = '   '
                        else:
                            cfsj_str = f"{st_dict['cfsj'].hour:2d}:"
                        cfsj_str += f'{st_dict["cfsj"].minute:02d}'
                        sec = st_dict['cfsj'].second
                        if sec:
                            ddsj_str += f"{sec:02d}"
                        else:
                            ddsj_str += '  '
                        ws.cell(row=row, column=col, value=cfsj_str)

                    else:
                        give_hour = False
                        if not last_dict:
                            give_hour = True
                        elif last_dict['cfsj'].hour != st_dict['ddsj'].hour:
                            give_hour = True
                        ws.cell(row=row + 1, column=col, value='   ...')
                        tgsj_str = f'{st_dict["ddsj"].hour:2d}:' if give_hour else '   '
                        tgsj_str += f'{st_dict["ddsj"].minute:02d}'
                        sec = st_dict['ddsj'].second
                        if sec:
                            tgsj_str += f"{sec:02d}"
                        else:
                            tgsj_str += '  '
                        ws.cell(row=row, column=col, value=tgsj_str)
                col += 1
                last_train = train

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center',
                                                                   vertical='center', shrink_to_fit=True)
                ws.cell(row=row, column=col).font = Font(name='宋体', size=9)

        wb.save(filename)

    def outTime(self, tgsj, give_hour: bool):
        tgsj_str = f'{tgsj.hour:2d}:' if give_hour else '   '
        tgsj_str += f'{tgsj.minute:02d}'
        sec = tgsj.second
        if sec:
            tgsj_str += f"{sec:02d}"
        else:
            tgsj_str += '  '
        return tgsj_str

    def getIntervalTrains(self, start, end, trainFilter, *, businessOnly=False, stoppedOnly=False):
        """
        返回某个区间办客车次列表。数据结构为list<dict>。
        //2.1版本修改逻辑为：两站皆办理业务才被选入。
        2019.06.29修改逻辑：选入的条件由输入参数给定。其中stoppedOnly包含始发终到情况。
        dict{
            'train':train object,
            'isSfz':boolean,
            'isZdz':boolean,
            'from':str,
            'to':str,
        """
        interval_list = []
        for train in self.trains():
            if not trainFilter.check(train):
                continue
            start_idx, end_idx = train.stationIndexByName(start), train.stationIndexByName(end)
            if start_idx == -1 or end_idx == -1:
                continue
            if start_idx > end_idx:
                continue
            start_dict, end_dict = train.timetable[start_idx], train.timetable[end_idx]

            if not (self.judgeStopAndBusiness(train, start_dict, businessOnly, stoppedOnly) and
                    self.judgeStopAndBusiness(train, end_dict, businessOnly, stoppedOnly)):
                continue
            isSfz = train.isSfz(start)
            isZdz = train.isZdz(end)
            train_dict = {
                'train': train,
                'isSfz': isSfz,
                'isZdz': isZdz,
                'from': start_dict['zhanming'],
                'to': end_dict['zhanming']
            }
            interval_list.append(train_dict)
        return interval_list

    def judgeStopAndBusiness(self, train: Train, dct: dict, bOnly: bool, sOnly: bool):
        """
        为上一个函数服务的工具性函数。判断时刻表中某车站是否符合对营业和停车的要求。
        表达式写的丑是为了利用短路性提高效率。
        """
        zm = dct['zhanming']
        return (not sOnly or train.stationStopped(dct) or train.isSfz(zm) or train.isZdz(zm)) and \
               (not bOnly or train.stationBusiness(dct))

    def getIntervalCount(self, fromOrTo, isStart, trainFilter, passenger_only=False, freight_only=False,
                         business_train_only=False, stopped_train_only=False):
        """
        获取区间对数表。
        :param fromOrTo:发站或到站
        :param isStart: True for start station, vice versa
        后两个参数：是否仅包括办客和办货的车站。中心站（fromOrTo）不受限制。
        返回数据结构list<dict>
        dict{
        'from'
        """
        infoList = []
        if isStart:
            for st in self.businessStationNames(passenger_only, freight_only):
                if not stationEqual(fromOrTo, st):
                    infoList.append({'from': fromOrTo, 'to': st, 'info':
                        self.getIntervalTrains(fromOrTo, st, trainFilter, businessOnly=business_train_only,
                                               stoppedOnly=stopped_train_only
                                               )})
        else:
            for st in self.businessStationNames(passenger_only, freight_only):
                if not stationEqual(fromOrTo, st):
                    infoList.append({'to': fromOrTo, 'from': st, 'info':
                        self.getIntervalTrains(st, fromOrTo, trainFilter, businessOnly=business_train_only,
                                               stoppedOnly=stopped_train_only
                                               )})

        count_list = []
        for info_dict in infoList:
            info = info_dict['info']
            count = len(tuple(info))
            countSfz = len([1 for st in info if st['isSfz']])
            countZdz = len([1 for st in info if st['isZdz']])
            countSfZd = len([1 for st in info if st['isZdz'] and st['isSfz']])
            int_dict = {
                'from': info_dict['from'],
                'to': info_dict['to'],
                'count': count,
                'countSfz': countSfz,
                'countZdz': countZdz,
                'countSfZd': countSfZd
            }
            count_list.append(int_dict)
        return count_list

    def getIntervalCount_faster(self, fromOrTo, isStart, trainFilter,
                                passenger_only=False, freight_only=False,
                                business_train_only=False, stopped_train_only=False) -> list:
        """
        2019.07.12新增，破坏封装性提高效率。
        原理是避免车次的时刻表被多次遍历。由于Line对象的name->dict有映射表而可以以近常量的效率完成，
        故使用反复的graph.stationDict代替反复的train.stationDict可显著提高效率。
        """
        # 统计单源点车站对数的四个表。数据结构为str,int，没有的站即为0.
        if not self.stationInLine(fromOrTo):
            return []
        startEndCount = {}
        startCount = {}
        endCount = {}
        allCount = {}
        if isStart:
            for train in self.trains():
                if not trainFilter.check(train):
                    continue
                started = 0  # 0: 未开始；1：开始但不是始发站；2：开始且是始发站。
                for st_dict_train in train.stationDicts():
                    st_dict_line = self.stationByDict(st_dict_train['zhanming'])
                    if st_dict_line is None:
                        continue
                    if stationEqual(st_dict_line['zhanming'], fromOrTo, strict=True):
                        if not self.judgeStopAndBusiness(train,
                                                         st_dict_train, business_train_only, stopped_train_only):
                            break
                        elif train.isSfz(st_dict_train['zhanming']):
                            started = 2
                        else:
                            started = 1
                        continue
                    if not started:
                        continue
                    # 到这里为止，保证站名存在，且已经越过了始发站。
                    zm = st_dict_line['zhanming']
                    # 排除不符合线路站点要求的项目
                    if not self.verifyStationBusiness(st_dict_line, passenger_only, freight_only):
                        continue
                    # 排除不符合车次营业及停车要求的项目
                    if not self.judgeStopAndBusiness(train, st_dict_train,
                                                     business_train_only, stopped_train_only):
                        continue
                    allCount[zm] = allCount.get(zm, 0) + 1
                    if train.isZdz(st_dict_train['zhanming']):
                        endCount[zm] = endCount.get(zm, 0) + 1
                    if started == 2:
                        startCount[zm] = startCount.get(zm, 0) + 1
                        if train.isZdz(zm):
                            startEndCount[zm] = startEndCount.get(zm, 0) + 1
            count_list = []
            for st_name in self.businessStationNames(passenger_only, freight_only):
                count_list.append(
                    {
                        'from': fromOrTo,
                        'to': st_name,
                        'count': allCount.get(st_name, 0),
                        'countSfz': startCount.get(st_name, 0),
                        'countZdz': endCount.get(st_name, 0),
                        'countSfZd': startEndCount.get(st_name, 0),
                    }
                )
            return count_list
        else:
            for train in self.trains():
                if not trainFilter.check(train):
                    continue
                started = 0  # 0: 未开始；1：开始但不是终到站；2：开始且是终到站。
                for st_dict_train in reversed(train.timetable):
                    st_dict_line = self.stationByDict(st_dict_train['zhanming'])
                    if st_dict_line is None:
                        continue
                    if stationEqual(st_dict_line['zhanming'], fromOrTo, strict=True):
                        if not self.judgeStopAndBusiness(train, st_dict_train, business_train_only,
                                                         stopped_train_only):
                            break
                        elif train.isZdz(st_dict_train['zhanming']):
                            started = 2
                        else:
                            started = 1
                        continue
                    if not started:
                        continue
                    # 到这里为止，保证站名存在，且已经越过了终到站。
                    zm = st_dict_line['zhanming']
                    # 排除不符合线路站点要求的项目
                    if not self.verifyStationBusiness(st_dict_line, passenger_only, freight_only):
                        continue
                    # 排除不符合车次营业及停车要求的项目
                    if not self.judgeStopAndBusiness(train, st_dict_train,
                                                     business_train_only, stopped_train_only):
                        continue
                    allCount[zm] = allCount.get(zm, 0) + 1
                    if train.isSfz(st_dict_train['zhanming']):
                        startCount[zm] = startCount.get(zm, 0) + 1
                    if started == 2:
                        endCount[zm] = endCount.get(zm, 0) + 1
                        if train.isSfz(zm):
                            startEndCount[zm] = startEndCount.get(zm, 0) + 1
            count_list = []
            for st_name in self.businessStationNames(passenger_only, freight_only):
                count_list.append(
                    {
                        'from': st_name,
                        'to': fromOrTo,
                        'count': allCount.get(st_name, 0),
                        'countSfz': startCount.get(st_name, 0),
                        'countZdz': endCount.get(st_name, 0),
                        'countSfZd': startEndCount.get(st_name, 0),
                    }
                )
            return count_list

    def stationByIndex(self, idx):
        return self.line.stations[idx]

    def resetAllTrainsLocalFirstLast(self):
        """
        当线路数据更新时，重置所有列车的localFirst/Last。
        """
        for train in self.trains():
            train.updateLocalFirst(self)
            train.updateLocalLast(self)

    def setMargin(self, ruler_label, mile_label, station_label, left_and_right,
                  top, bottom, system=False) -> bool:
        """
        1.4版本修改
        用户设置页边距。注意参数含义与本系统内部使用的不同。返回是否变化。
        """
        left_white = 15
        right_white = 10
        margins = {
            "left_white": left_white,
            "right_white": right_white,
            "left": ruler_label + mile_label + station_label + left_and_right + left_white,
            "up": top,
            "down": bottom,
            "right": left_and_right + station_label + right_white,
            "label_width": station_label,
            "mile_label_width": mile_label,
            "ruler_label_width": ruler_label,
        }
        UIDict = self.UIConfigData() if not system else self.sysConfigData()
        changed = UIDict.get('margins', None) != margins
        UIDict["margins"] = margins
        return changed

    def toTrc(self, filename):
        fp = open(filename, 'w', encoding='utf-8', errors='ignore')
        fp.write('***Circuit***\n')
        fp.write(f"{self.lineName() if self.lineName() else '列车运行图'}\n")
        fp.write(f"{int(self.lineLength())}\n")
        last_dct = None
        for dct in self.stationDicts():
            fp.write(f"{dct['zhanming']},{int(dct['licheng'])},{dct['dengji']},"
                     f"{'false' if dct.get('show',True) else 'true'},,true,4,1440,1440,")
            if last_dct is not None:
                node = self.line.forbid.getInfo(last_dct['zhanming'], dct['zhanming'])
                if node is not None and node['begin'] != node['end']:
                    fp.write(f"{node['begin'].strftime('%H:%M')}-{node['end'].strftime('%H:%M')}\n")
                else:
                    fp.write('\n')
            else:
                fp.write('\n')
            last_dct = dct
        for train in self.trains():
            fp.write('===Train===\n')
            fp.write(f"trf2,{train.fullCheci()},{train.downCheci()},{train.upCheci()}")
            circuit = train.carriageCircuit()
            if isinstance(circuit, Circuit):
                fp.write(f",{circuit.name().replace('_','-')}_{circuit.trainOrderNum(train)}\n")
            else:
                fp.write(',NA\n')
            fp.write(f"{train.sfz if train.sfz else 'null'}\n")
            fp.write(f"{train.zdz if train.zdz else 'null'}\n")
            for name, ddsj, cfsj in train.station_infos():
                fp.write(f"{name},{ddsj.strftime('%H:%M:%S')},{cfsj.strftime('%H:%M:%S')},true,NA\n")
        fp.write('---Color---\n')
        for train in self.trains():
            color_str = train.color(self)
            fp.write(f"{train.fullCheci()},{int(color_str[1:3],base=16)},"
                     f"{int(color_str[3:5],base=16)},{int(color_str[5:7],base=16)}\n")
        fp.write('---LineType---\n')
        for train in self.trains():
            if train.lineWidth() != 0:
                fp.write(f"{train.fullCheci()},0,{train.lineWidth()}")
        fp.close()

    def circuitByName(self, name: str, *, throwError=True) -> Circuit:
        """
        根据交路名查找交路对象。如果没有对应交路，抛出CircuitNotFoundError。
        """
        for circuit in self._circuits:
            if circuit.name() == name:
                return circuit
        if throwError:
            raise CircuitNotFoundError(name)
        else:
            return None

    def circuits(self):
        for c in self._circuits:
            yield c

    def circuitCount(self):
        return len(self._circuits)

    def addCircuit(self, circuit: Circuit):
        """
        若名称已存在，抛出CircuitExistedError
        """
        if self.circuitByName(circuit.name(), throwError=False) is not None:
            raise CircuitExistedError(circuit.name())
        self._circuits.append(circuit)

    def delCircuit(self, circuit: Circuit):
        """
        若不存在，抛出CircuitNotFoundError
        """
        assert isinstance(circuit, Circuit)
        try:
            self._circuits.remove(circuit)
        except ValueError:
            raise CircuitNotFoundError(circuit.name())
        for node in circuit.nodes():
            if not node.isVirtual():
                try:
                    node.train().setCarriageCircuit(None)
                except AttributeError:
                    print("Graph::delCircuit: Unexcpeted node.train", node)
                except TrainNotFoundException as e:
                    print("Graph::delCircuit: TrainNotFoundException", e)
                    pass

    def checkGraph(self) -> str:
        """
        对运行图文件可能出现的问题做启动时的检查。返回是报错的字符串。没有其他影响。
        如果没有问题，返回空串。
        """
        report = ""
        # 检查正则表达式的问题
        i = 0
        found = False
        for name, rg, _ in self.UIConfigData()['type_regex']:
            try:
                re.compile(rg)
            except:
                if rg == r'7\d+{3}':
                    self.UIConfigData()['type_regex'][i] = (name, r'7\d{3}', _)
                else:
                    found = True
                    report += f"列车类型[{name}]的正则表达式[{rg}]错误！\n"
            else:
                pass
            i += 1
        if found:
            report += "请至“运行图设置”（ctrl+G）面板中“类型管理”中修正错误的正则表达式。\n"

        if report:
            report += "此信息提示当前运行图文件可能不符合此版本软件的要求。您可以忽略此提示，但有可能导致" \
                      "程序运行异常。下次打开本文件时，如果仍未解决，此消息仍会显示。\n"
        return report

    def modelList(self) -> list:
        """
        2019.06.25新增。返回所有不同的车底类型列表。
        目前只由trainFilter中筛选的有关功能调用，简化起见，不作为属性保存，每次都遍历得到。
        """
        lst = []
        for circuit in self.circuits():
            lst.append(circuit.model())
        return list(set(lst))

    def ownerList(self) -> list:
        """
        返回所有不同的担当局段表。
        """
        lst = []
        for circuit in self.circuits():
            lst.append(circuit.owner())
        return list(set(lst))

    def clearAll(self):
        """
        清空所有数据
        """
        self.clearLineStationInfo()
        self.setOrdinateRuler(None)
        self._trains = []
        self._circuits = []
        self.typeList = []
        self.fullCheciMap = {}
        self.singleCheciMap = {}
        self._markdown = ''
        self.filename = ''

    def clearTrains(self):
        self._trains = []
        self._circuits = []
        self.typeList = []
        self.fullCheciMap = {}
        self.singleCheciMap = {}

    def nameMapToLine(self, name: str):
        """
        支持域解析符的情况下，将车次中的站名映射到本线站名。
        """
        dct = self.stationByDict(name)
        try:
            return dct['zhanming']
        except TypeError:
            return name

    class TrainDiffType(Enum):
        Unchanged = 0
        Changed = 1
        NewAdded = 2
        Deleted = 3

    def diffWith(self, graph, callBack=None) -> List[Tuple[
        TrainDiffType,
        Union[List[Tuple[
            Train.StationDiffType, TrainStation, TrainStation
        ]], None],
        Union[int, None],
        Union[Train, None],
        Union[Train, None]
    ]]:
        """
        与graph所示运行图进行车次比较。返回：与车次时刻表对比类似：List<Tuple>。但是新增一项，不同的数目。
        ( 类型，车次比较报告, 不同数目，train1, train2,)
        如果某个列车不存在则该列车train位置, 车次比较报告位置，和不同的数目位置都是None。
        不考虑车次的顺序。直接利用对方的fullCheciMap，浅拷贝一次。
        前置条件：车次是identical的。
        2019.11.17修改：将判断是否是本线车的逻辑移到Dialog里面。理由是减少反复的比较，而仅在读取的时候比较一次。判断是否本线车的代价要远小于比较交叉车次的代价，故此。
        """
        tm1 = time.perf_counter()
        graph: Graph
        result = []
        anotherFullMap = graph.fullCheciMap.copy()
        for train in self.trains():
            checi = train.fullCheci()
            train2 = anotherFullMap.get(checi, None)
            if train2 is None:
                result.append((
                    Graph.TrainDiffType.Deleted, None, None, train, None
                ))
                if callBack:
                    callBack(1)
            else:
                trainDiffData, value = train.globalDiff(train2)  # 最慢
                tp = Graph.TrainDiffType.Unchanged if value == 0 else Graph.TrainDiffType.Changed
                result.append((
                    tp, trainDiffData, value, train, train2
                ))
                del anotherFullMap[checi]
                if callBack:
                    callBack(2)
        for _, train2 in anotherFullMap.items():
            train2: Train
            result.append((
                Graph.TrainDiffType.NewAdded, None, None, None, train2
            ))
            if callBack:
                callBack(1)
        tm2 = time.perf_counter()
        print("Graph::diffWith()历时", tm2 - tm1)
        return result

    def subGraph(self,line:Line,withAllTrains=False):
        """
        给出指定line下的子图。
        """
        graph = Graph()
        graph.setLine(line)
        # 先导入所有交路
        for circuit in self.circuits():
            c = Circuit(graph)
            c.coverBaseData(circuit)
            graph.addCircuit(c)
        for train in self.trains():
            if withAllTrains or train.isLocalTrain(graph):
                circuit = train.carriageCircuit()
                if circuit is not None:
                    train.setCarriageCircuit(graph.circuitByName(circuit.name()))
                graph.addTrain(train)
        for circuit in graph._circuits[:]:
            if not circuit.anyValidTrains():
                graph._circuits.remove(circuit)
        return graph

    def rulerFromMultiTrains(self, intervals:List[Tuple[str,str]], trains:List[Train],
                             different:bool, useAverage:bool,
                             defaultStart:int, defaultStop:int,
                             cutSigma:int=None, cutSeconds:int=None,
                             prec:int=1,
                             cutCount:int=1,
                             )->(Dict[Tuple[str,str],Tuple[int,int,int]],
                                 Dict[Tuple[str, str], Dict[Train, Tuple[int, int]]],
                                 Dict[Tuple[str, str], Dict[int, Dict[int, int]]],
                                 Dict[Tuple[str, str], Dict[int, Tuple[int, int, bool]]]):
        """
        2020.03.13新增。从一组给定的（并假定拥有相同标尺）的车次中读取标尺。
        假定各个车次各个区间的运行情况是独立的；即不认为一个车次各个区间的标尺是相同的。
        取区间数据的方法和Ruler.rulerFromTrain相同。
        :param different 上下行是否分设。如果不分设，上下行车次同时考虑。
        :param useAverage 用平均数计算，或者用众数计算。
        :param defaultStart 如果缺失数据，默认情况下用这个数据为起步附加时分。
        :param cutSigma 用平均数计算时，去除超过标准差此倍数的数据。
        :param cutSeconds 用平均数计算时，去除与平均数相差超过此数值的数据。
        如果同时设置，cutSeconds优先（计算量小一些）。
        :param prec 结果保留精度，单位为秒。众数时，如果最多的有多个数，也要取平均数到这个精度。
        :param cutCount 最低采纳的数据类。用于解决只有1条数据的异常数据类污染整个数据集问题。
        :returns (标尺数据，列车数据打表, FT数据打表, 最终采用数据表)
        最终采信数据三元组：[各区间、类型的采信数据，数据量，是否代入]
        """
        intSet = set(intervals)  # 用集合实现速查。对车次的每个站名，必须映射到本线，通过stationDictByName

        # 对每个区间和车次打表。数据结构：
        # dict (interval => dict (train => data, 区间附加标记))
        data = {}  # type:Dict[Tuple[str,str],Dict[Train,Tuple[int,int]]]
        for train in trains:
            lastSt:TrainStation = None
            lastLineSt:LineStation = None
            for st in train.stationDicts():
                line_dct = self.stationByDict(st['zhanming'])
                if line_dct is None:
                    continue
                # 本线第一个站的情况
                if lastSt is None:
                    lastSt = st
                    lastLineSt = line_dct
                    continue
                # 现在保证上站和本站都存在，且是本线区间
                tupInt = (lastLineSt['zhanming'],line_dct['zhanming'])
                if tupInt in intSet or (not different and
                        (line_dct['zhanming'],lastLineSt['zhanming']) in intSet):
                    # 属于要查找的区间
                    data.setdefault(tupInt,{})[train] = (
                        Train.dt(lastSt['cfsj'],st['ddsj']), train.intervalAttachType(lastSt,st)
                    )

                # continuing loop
                lastSt = st
                lastLineSt = line_dct

        res = {}
        ft = {}
        used = {}
        for (fazhan, daozhan), int_dct in data.items():
            int_data_trans = self.__intervalFt(int_dct)
            ft[(fazhan,daozhan)] = int_data_trans
            if useAverage:
                res[(fazhan,daozhan)],used[(fazhan,daozhan)] = self.__intervalRulerMean(int_data_trans,
                                                                                        defaultStart,
                                                               defaultStop,prec,cutSigma,cutSeconds,cutCount)
            else:  # 众数模式
                res[(fazhan,daozhan)],used[(fazhan,daozhan)] = self.__intervalRulerMode(int_data_trans,defaultStart,defaultStop,prec,cutCount)
        return res, data, ft, used

    def __intervalFt(self,dct:Dict[Train,Tuple[int,int]])->Dict[int,Dict[int,int]]:
        """
        对车次区间各种起停附加的情况的各种数据分别统计频数。类似Fourier transform
        """
        res = {}
        for _,(sec, tp) in dct.items():
            d = res.setdefault(tp,{})
            d[sec]=d.get(sec,0)+1
        return res

    def __intervalRulerMode(self, data:Dict[int,Dict[int,int]], defaultStart:int,
                            defaultStop:int,prec:int,cutCount:int,
                            )->(int,int,int,Dict[int,Tuple[int,int,bool]]):
        """
        众数模式计算给定区间的标尺。
        众数模式下，如果能删除一个数据（最少的一个数据与次少的数量不一致），则不适用伪逆。
        :returns ((interval, start, stop), 各类型使用数据的报告)
        """
        from copy import deepcopy
        modes = {}  # int->int  类型及其众数
        used = {}
        for tp, count_dct in data.items():  # count_dict: int->int  数据及其计数
            # 保证count_dct不是空的
            lst = list(sorted(count_dct.items(),key=lambda x:(x[1],-x[0]),reverse=True))
            modes[tp] = lst[0][0]  # 相同时，取快的那个
            used[tp] = (lst[0][0], lst[0][1], True)
        # 2020.08.23添加  删除小于指定有效数据量的数据类
        for tp,(val,cnt,_) in used.copy().items():
            if cnt < cutCount:
                del modes[tp]
                used[tp] = (val,cnt,False)
        if len(modes) == 4:
            # 先找出每一类的第一个数据及其出现次数
            ls = list(sorted(list(map(lambda x:(x[0],max(x[1].values())),data.items())),
                             key=lambda x:x[1]))
            if ls[0][1] != ls[1][1]:
                del modes[ls[0][0]]
                used[ls[0][0]] = (used[ls[0][0]][0],used[ls[0][0]][1],False)
            else:
                # 选择计算结果最小的那一个
                i = 1
                while i<len(ls) and ls[i][1] == ls[0][1]:
                    i += 1
                # i标记最后一个相同的
                res_list = []
                for t in range(i):
                    # 尝试剔除掉这个，然后算一个结果出来
                    modes_cpy = deepcopy(modes)
                    del modes_cpy[ls[t][0]]
                    res_list.append((self.__computeIntervalRuler(modes_cpy,defaultStart,defaultStop,prec),t))
                res,n = min(res_list)
                del modes[ls[n][0]]
                used[ls[n][0]] = (used[ls[n][0]][0],used[ls[n][0]][1],False)
                return res,used
        return self.__computeIntervalRuler(modes,defaultStart,defaultStop,prec),used

    def __intervalRulerMean(self, data:Dict[int,Dict[int,int]], defaultStart:int,
                            defaultStop:int,prec:int,
                            cutSigma:int=None,
                            cutSeconds:int=None,cutCount:int=None
                            )->(Tuple[int,int,int],Dict[int,Tuple[int,int,bool]]):
        """
        平均数模式计算给定区间的标尺。
        """

        def moment(lst:List[Tuple[int, int]])->(float,float):
            """返回均值和样本标准差"""
            N = sum(map(lambda x:x[1],lst))  # 样本总量, 保证大于1
            if N == 1:
                return lst[0][0],0
            ave = reduce(lambda x,y:x+y[0]*y[1],lst,0)/N
            sigma = sqrt(reduce(lambda x,y:x+(y[0]-ave)**2*y[1],lst,0)/(N-1))
            return ave, sigma

        def furthest(lst, av)->int:
            """返回距离均值最远的那个index，自然是第一个或最后一个。保证输入非空"""
            if abs(lst[0][0]-av) > abs(lst[-1][0]-av):
                return 0
            return -1

        means = {}
        used = {}  # 均值情况的数据量为剩余数据总量
        for tp,count_dct in data.items():
            # 按值排序，每次选择距离中心最远的一个决定要不要删掉。
            lst:List[Tuple[int,int]] = list(sorted(count_dct.items(),reverse=True))
            # 修约数据。删除偏差太大的数据。
            if cutSeconds:
                while len(lst) > 1:
                    ave,sigma = moment(lst)
                    i = furthest(lst,ave)
                    if abs(lst[i][0]-ave) > cutSeconds:
                        value,_ = lst.pop(i)
                        del count_dct[value]
                    else:
                        break
            elif cutSigma:
                while len(lst) > 1:
                    ave,sigma = moment(lst)
                    i = furthest(lst,ave)
                    if sigma and abs(lst[i][0]-ave)/sigma > cutSigma:
                        value,_ = lst.pop(i)
                        del count_dct[value]
                    else:
                        break
            ave,_ = moment(lst)
            means[tp] = ave
            used[tp] = (ave, sum(count_dct.values()), True)

        # 2020.08.23新增 清洗完数据后，最后检查数据量是否符合要求。
        for tp,(ave,cnt,_) in used.copy().items():
            if cnt < cutCount:
                del means[tp]
                used[tp] = (ave, cnt, False)

        return self.__computeIntervalRuler(means,defaultStart,defaultStop,prec), used

    @staticmethod
    def __round(value:int, prec:int)->int:
        """
        按“四舍五入”原则将value修约到prec为单位的数值。
        :return invariant: ret % prec == 0
        """
        if value < 0:
            value = 0
        q = value % prec
        if q == 0:
            return value
        elif q >= prec/2:
            return value+prec-q
        else:
            return value-q

    @staticmethod
    def __computeIntervalRuler(values:Dict[int,int],defaultStart:int,defaultStop:int,
                               prec:int)->(int,int,int):
        """
        由算好的数据计算所给区间的区间数据，起步附加，停车附加
        :param values 非空字典
        根据自由度数分析。
        """
        a, b, c, d = (values.get(Train.AttachNone), values.get(Train.AttachStart),
                      values.get(Train.AttachStop), values.get(Train.AttachBoth))
        # 多了个数据，用伪逆矩阵求解
        if len(values) == 4:
            x,y,z = (
                0.75*a+0.25*b+0.25*c-0.25*d,
                -0.5*a+0.50*b-0.50*c+0.50*d,
                -0.5*a-0.50*b+0.50*c+0.50*d
            )
        elif len(values) == 3:
            # 3个数据，刚好有唯一解。打表
            if a is None:
                x,y,z = (b+c-d,-c+d,-b+d)
            elif b is None:
                x,y,z = (a, -c+d, -a+c)
            elif c is None:
                x,y,z = (a,-a+b,-b+d)
            else:  # d is None
                x,y,z = (a,-a+b,-a+c)
        elif len(values) == 2:
            if a is not None:
                if b is not None:
                    x,y,z = (a,b-a,defaultStop)
                elif c is not None:
                    x,y,z = (a,defaultStart,c-a)
                else:  # d is not None:
                    x,y,z = (a,defaultStart,d-a-defaultStart)
            else:  # a is None
                if b is None:  # c,d
                    x,y,z = (c-defaultStop,d-c,defaultStop)
                elif c is None:  # b,d
                    x,y,z = (b-defaultStart,defaultStart,d-b)
                else:  # d is None, b,c
                    x,y,z = (b-defaultStart,defaultStart,c-b+defaultStart)
        else:  # 1个自由度
            y,z = defaultStart,defaultStop
            if a is not None:
                x = a
            elif b is not None:
                x = b-defaultStart
            elif c is not None:
                x = c-defaultStop
            else:  # d
                x = d-defaultStart-defaultStop
        return Graph.__round(x,prec),Graph.__round(y,prec),Graph.__round(z,prec)

    def importTrainFromExcel(self,filename)->(int,int,int):
        """
        从指定Excel文件中导入时刻表数据。返回：
        导入成功数据数，失败数据数，新建车次数
        """
        import xlrd
        suc,fail,new = 0,0,0
        wb = xlrd.open_workbook(filename)
        for ws in wb.sheets():
            for row in range(ws.nrows):
                try:
                    cc = ws.cell_value(row,0)
                    zm = ws.cell_value(row,1)
                    dd = ws.cell_value(row,2)
                    if isinstance(dd,str):
                        dd = strToTime(dd)
                    else:
                        dd = xlrd.xldate_as_datetime(dd,1)
                    cf = ws.cell_value(row,3)
                    if isinstance(cf,str):
                        cf = strToTime(cf)
                    else:
                        cf = xlrd.xldate_as_datetime(cf,1)
                    try:
                        track = ws.cell_value(row,4)
                    except IndexError:
                        track = ''
                    try:
                        note = ws.cell_value(row,5)
                    except IndexError:
                        note = ''
                    train = self.trainFromCheci(cc,full_only=True)
                    if train is None:
                        train = Train(self,checi_full=cc)
                        self.addTrain(train)
                        new+=1
                    train.addStation(zm,dd,cf,note=note,track=track)
                    suc+=1
                except Exception as e:
                    print(f"Graph::importTrainFromExcel: import failed. "
                          f"Sheet: {ws.name}, row: {row+1}"+repr(e))
                    fail+=1
        return suc,fail,new

    def updateTypeList(self):
        lst = []
        self.typeList.clear()
        for train in self.trains():
            lst.append(train.trainType())
        self.typeList.extend(set(lst))

if __name__ == '__main__':
    graph = Graph()
    graph.loadGraph("source/output.json")
    graph.show()
    graph.save("source/test.json")
