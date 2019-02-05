"""
运行图类
"""
from line import Line
from ruler import Ruler
from train import Train
from copy import copy
from utility import stationEqual
import json

config_file = 'config.json'

import cgitb
cgitb.enable(format='text')

class Graph():
    """
    运行图类，数据结构：
    Line line;
    List<Train> _trains=[];
    List<Dict> _circuit;//交路数据
    Dict _config;//系统设置，主要是UI
    str markdown;//附注
    交路格式：
    Dict{
        "name":交路名称,
        "trains":['G89','G90',...]
    """
    def __init__(self):
        """
        构造空类，不考虑读文件
        """
        self.filename = ""
        self.line = Line()
        self._trains = []
        self._circuits = []
        self._config = {}
        self.typeList = [] # public data
        self.readConfig()
        self._markdown = ''
        self.fullCheciMap = {} # 全车次查找表 str|->Train
        self.singleCheciMap = {} # 单车次查找表str|->list<Train>

    def setFullCheciMap(self):
        """
        设置全车次查找表。
        """
        for train in self.trains():
            self.fullCheciMap[train.fullCheci()]=train

    def setSingleCheciMap(self):
        for train in self.trains():
            for cc in (train.downCheci(),train.upCheci()):
                if cc:
                    self.singleCheciMap.setdefault(cc,[]).append(train)

    def addSingleCheciMap(self,train):
        for cc in (train.downCheci(),train.upCheci()):
            if cc:
                self.singleCheciMap.setdefault(cc,[]).append(train)

    def delSingleCheciMap(self,train):
        for cc in (train.downCheci(),train.upCheci()):
            if cc:
                lst = self.singleCheciMap.get(cc)
                if len(lst) > 1:
                    lst.remove(train)
                else:
                    del self.singleCheciMap[cc]

    def readConfig(self):
        fp = open(config_file,encoding='utf-8',errors='ignore')
        buff = json.load(fp)
        if buff["ordinate"] is not None:
            buff["ordinate"] = self.line.rulerByName(buff["ordinate"])
        fp.close()
        self._config=buff

    def readSystemConfig(self):
        fp = open(config_file, encoding='utf-8', errors='ignore')
        buff = json.load(fp)
        return buff

    def saveConfig(self):
        print("=============saveConfig")
        print(self._config)
        if self._config is None or not self._config:
            raise Exception("Cannot save empty config data")
        data = copy(self._config)
        try:
            data["ordinate"]
        except KeyError:
            pass
        else:
            if data["ordinate"] is not None:
                data["ordinate"] = data["ordinate"].name()

        fp = open(config_file,'w',encoding='utf-8',errors='ignore')
        json.dump(data,fp,ensure_ascii=False)
        fp.close()

    def loadGraph(self,filename:str):
        """
        暂定直接打开json文件读
        """
        self.filename = filename
        fp = open(filename,encoding='utf8',errors='ignore')
        try:
            info = json.load(fp)
        except json.JSONDecodeError:
            fp.close()
            self.loadTrcGraph(filename)
            return

        self.line.loadLine(info["line"])
        self.circuits = info["circuits"]
        for dict_train in info["trains"]:
            newtrain = Train(origin=dict_train)
            self._trains.append(newtrain)
        try:
            self._config = info["config"]
            try:
                if self._config["ordinate"] is not None:
                    self._config["ordinate"] = self.line.rulerByName(self._config["ordinate"])
            except:
                pass
        except KeyError:
            self.readConfig()
        else:
            if self._config is None or not self._config:
                self.readConfig()

        try:
            self._markdown = info["markdown"]
        except KeyError:
            self._markdown = ''

        fp.close()
        self.setFullCheciMap()
        self.setSingleCheciMap()

    def addCircuit(self):
        pass

    def addTrain(self,train:Train):
        self._trains.append(train)
        self.fullCheciMap[train.fullCheci()] = train
        self.addSingleCheciMap(train)

    def delTrain(self,train:Train):
        try:
            self._trains.remove(train)
            del self.fullCheciMap[train.fullCheci()]
            self.delSingleCheciMap(train)
        except:
            print("del train: No such train!",train.fullCheci())

    def trains(self):
        for train in self._trains:
            yield train

    def show(self):
        self.line.show()
        print(self.circuits)
        for train in self._trains:
            train.show()

    def setLine(self,line:Line):
        self.line = line

    def lineLength(self):
        try:
            return self.line.stations[-1]["licheng"]
        except IndexError:
            return 0

    def stations(self,reverse=False):
        if not reverse:
            for station in self.line.stations:
                yield station["zhanming"]
        else:
            for station in reversed(self.line.stations):
                yield station["zhanming"]

    def stationDicts(self,reverse=False):
        if not reverse:
            for station in self.line.stations:
                yield station
        else:
            for station in reversed(self.line.stations):
                yield station

    def save(self,filename:str):
        """
        保存运行图文件
        """
        graph = {
            "line":self.line.outInfo(),
            "trains":[],
            "circuits":self._circuits,
            "config":self._config,
        }
        try:
            graph["markdown"] = self._markdown
        except AttributeError:
            self._markdown = ''
            graph["markdown"] = ''
        if graph["config"]["ordinate"] is not None:
            if isinstance(graph["config"]["ordinate"],str):
                pass
            else:
                graph["config"]["ordinate"] = graph["config"]["ordinate"].name()

        for train in self._trains:
            graph["trains"].append(train.outInfo())

        with open(filename,'w', encoding='utf8', errors='ignore') as fp:
            # print(graph["line"]["rulers"])
            json.dump(graph,fp,ensure_ascii=False)

    def UIConfigData(self):
        return self._config

    def line_station_mileages(self):
        """
        生成器，返回 dict["zhanming"],dict["licheng"]
        """
        for dict in self.line.stations:
            yield dict["zhanming"],dict["licheng"]

    def line_station_mile_levels(self):
        for dict in self.line.stations:
            yield dict["zhanming"],dict["licheng"],dict["dengji"]

    def addEmptyRuler(self,name:str,different:bool=False):
        ruler = Ruler(name=name,different=different,line=self.line)
        self.line.rulers.append(ruler)
        return ruler

    def addRuler(self,ruler:Ruler):
        self.line.rulers.append(ruler)

    def delRuler(self,ruler:Ruler):
        """
        返回被删除的标尺是否是排图标尺
        """
        try:
            self.line.rulers.remove(ruler)
        except ValueError:
            pass
        if self.ordinateRuler() is ruler:
            self.setOrdinateRuler(None)
            return True
        return False

    def setOrdinateRuler(self,ruler:Ruler):
        self._config["ordinate"] = ruler

    def ordinateRuler(self):
        try:
            return self._config["ordinate"]
        except KeyError:
            self._config["ordinate"] = None
            return None

    def setStationYValue(self,name,y):
        """
        设置某个站的纵坐标值。2019.02.02移植到line类并删除线性算法。
        """
        self.line.setStationYValue(name,y)

    def stationYValue(self,name:str):
        st_dict = self.stationByDict(name)
        if st_dict is not None:
            return st_dict.get('y_value',-1)
        else:
            return -1

    def stationMileYValues(self):
        for dct in self.line.stations:
            yield dct['zhanming'],dct['licheng'],dct.get('y_value',None)

    def trainFromCheci(self,checi:str,full_only=False):
        """
        根据车次查找Train对象。如果full_only，则仅根据全车次查找；否则返回单车次匹配的第一个结果。
        若不存在，返回None。
        2019.02.03删除线性算法。
        """
        t = self.fullCheciMap.get(checi,None)
        if t is not None:
            return t
        if not full_only:
            selected = self.singleCheciMap.get(checi,None)
            if selected is not None:
                return selected[0]
        return None

    def multiSearch(self,checi:str):
        """
        非严格搜索。线性算法。
        """
        selected = []
        for train in self.trains():
            if checi in train.fullCheci() or checi in train.downCheci() or checi in train.upCheci():
                selected.append(train)
        return selected

    def lineName(self):
        return self.line.name

    def setLineName(self,name:str):
        self.line.name = name

    def firstStation(self):
        return self.line.stations[0]["zhanming"]

    def lastStation(self):
        return self.line.stations[-1]["zhanming"]

    def stationInLine(self,name:str,strict=False):
        return self.line.stationInLine(name,strict)

    def setStationIsShow(self,name:str,show:bool):
        """
        不支持域解析符。2019.02.02删除线性算法。
        """
        st = self.line.stationDictByName(name,True)
        if st is None:
            raise Exception("setStationIsShow: no such station")
        else:
            st['show'] = show

    def stationIsShow(self,name:str):
        dct = self.line.stationDictByName(name)
        if dct is None:
            raise Exception("No such station")
        return dct.setdefault('show',True)

    def graphFileName(self):
        return self.filename

    def setGraphFileName(self,filename:str):
        self.filename=filename

    def isEmpty(self):
        if self.line.stations:
            return False
        else:
            return True

    def stationCount(self):
        return len(self.line.stations)

    def clearLineStationInfo(self):
        del self.line.stations
        self.line.stations = []

    def stationExisted(self,name:str):
        self.line.stationExisted(name)

    def addStationDict(self,info:dict):
        self.line.stations.append(info)

    def adjustLichengTo0(self):
        if self.isEmpty():
            return
        start_mile = self.line.stations[0]["licheng"]
        for st in self.line.stations:
            st["licheng"] = st["licheng"]-start_mile

    def trainCount(self):
        return len(self._trains)

    def rulers(self):
        for ruler in self.line.rulers:
            yield ruler

    def gapBetween(self,st1:str,st2:str):
        """
        计算两个站间距离
        """
        station1 = self.stationByDict(st1)
        station2 = self.stationByDict(st2)

        if station1 is None or station2 is None:
            raise Exception("No such station",st1,st2)

        return abs(station1["licheng"]-station2["licheng"])

    def lineSplited(self):
        """
        返回本线是否存在上下行分设站的情况
        :return:
        """
        if self.line.isSplited():
            return True
        return False

    def rulerNameExisted(self,name,ignore:Ruler=None):
        for r in self.line.rulers:
            if r is not ignore and r.name() == name:
                return True
        return False

    def isNewRuler(self,ruler:Ruler):
        for r in self.line.rulers:
            if ruler is r:
                return False
        return True

    def stationDirection(self,name:str):
        return self.line.stationViaDirection(name)

    def formerBothStation(self,name:str):
        """
        寻找本站往前的第一个【下行方向通过】的站。
        TODO 2019.02.02 保留线性算法。下一个函数同。
        """
        former_dict = None
        for st in self.line.stations:
            if st["zhanming"] == name:
                return former_dict

            if st["direction"] == 0x3:
                former_dict = st
        raise Exception("No former station")

    def latterBothStation(self,name:str):
        start = False
        for st in self.line.stations:
            if st["zhanming"] == name:
                start = True
            if start and st["direction"] == 0x3:
                return st
        raise Exception("No latter station")

    def stationLevel(self,name:str):
        """
        返回车站等级。若不存在，返回None；若没有这个字段，设为并返回4. 不支持域解析符
        """
        st = self.line.stationDictByName(name,strict=True)
        if st is None:
            return None
        return st.setdefault('dengji',4)

    def setNotShowTypes(self,not_show):
        self.UIConfigData()["not_show_types"] = not_show
        for train in self.trains():
            if train.type in not_show:
                train.setIsShow(False,affect_item=False)
            else:
                train.setIsShow(True,affect_item=False)

    def setDirShow(self,down,show):
        for train in self.trains():
            if train.isDown() == down:
                train.setIsShow(show,affect_item=False)

    def trainExisted(self,train:Train,ignore:Train=None):
        """
        比较Train对象，线性算法
        """
        for t in self._trains:
            if train is t and t is not ignore:
                return True
        return False

    def checiExisted(self,checi:str,ignore:Train=None):
        """
        比较全车次。2019.02.03替换掉线性算法。
        """
        # for t in self._trains:
        #     if t is not ignore and t.fullCheci() == checi:
        #         return True
        # return False
        t = self.fullCheciMap.get(checi,None)
        if t is not None and t is not ignore:
            return True
        return False

    def rulerCount(self):
        return len(self.line.rulers)

    def stationTimeTable(self,name:str):
        """
        返回车站的图定时刻表
        list<dict>
        dict{
            "station_name":str,
            "ddsj":datetime,
            "cfsj":datetime,
            "down":bool,
            "train":Train,
        }
        """
        timeTable = []
        for train in self.trains():
            st_dict = train.stationDict(name)
            if st_dict is None:
                continue
            else:
                node = {
                    "ddsj":st_dict["ddsj"],
                    "cfsj":st_dict["cfsj"],
                    "station_name":st_dict["zhanming"],
                    "down":train.isDown(auto_guess=True,graph=self),
                    "train":train,
                }
                timeTable.append(node)

        #排序
        for i in range(len(timeTable)-1):
            t = i
            for j in range(i+1,len(timeTable)):
                if timeTable[j]["ddsj"] < timeTable[t]["ddsj"]:
                    t = j
            temp = timeTable[t];timeTable[t] = timeTable[i];timeTable[i]=temp
        return timeTable

    def reverse(self):
        """
        反排运行图
        :return:
        """
        length = self.lineLength()

        #里程调整
        for st_dict in self.line.stations:
            st_dict["licheng"] = length - st_dict["licheng"]
        self.line.stations.reverse()

        #列车上下行调整、上下行车次交换
        for train in self._trains:
            #上下行交换
            if train.isDown() is not None:
                train.setIsDown(not train.isDown())

            #车次交换
            temp = train.setCheci(train.fullCheci(),train.upCheci(),train.downCheci())

    def downTrainCount(self):
        count = 0
        for train in self.trains():
            if train.isDown() is True:
                count+=1
        return count

    def upTrainCount(self):
        count = 0
        for train in self.trains():
            if train.isDown() is False:
                count+=1
        return count

    def loadTrcGraph(self,filename):
        """
        阅读旧版trc格式的运行图
        """
        fp = open(filename,encoding='utf-8',errors='ignore')

        inTrainArea = False
        now_list = []
        for i,line in enumerate(fp):
            line = line.strip()
            if not line:
                continue
            if not inTrainArea and line == "===Train===":
                inTrainArea=True

            if line[0] == '-':
                break

            #处理线路信息部分
            if not inTrainArea:
                if line == "***Circuit***":
                    continue
                elif i == 1:
                    self.setLineName(line)
                else:
                    #线路信息部分
                    try:
                        splited = line.split(',')
                        self.line.addStation_by_info(splited[0],int(splited[1]),int(splited[2]))
                    except:
                        pass

            #处理列车信息部分
            else:
                #这部分从trc_check_new中复制过来
                if line != '===Train===':
                    now_list.append(line)
                else:
                    self._decodeTrcTrain(now_list)
                    now_list = []
        self._decodeTrcTrain(now_list)
        self.setGraphFileName('')

    def _decodeTrcTrain(self,now_list:list):
        """
        阅读trc中单个车次的信息，不含===Train===标志头
        """
        train = Train()
        for i,line in enumerate(now_list):
            if i == 0:
                splited = line.split(',')
                train.setCheci(splited[1],splited[2],splited[3])

            elif i == 1:
                train.setStartEnd(sfz=line)
            elif i == 2:
                train.setStartEnd(zdz=line)
            else:
                splited = line.split(',')
                train.addStation(splited[0],splited[1],splited[2])
        train.autoType()
        if train.timetable:
            self.addTrain(train)

    def jointGraph(self,graph,former:bool,reverse:bool,line_only:bool):
        """
        拼接两运行图。
        :param graph: 另一运行图
        :param former: 另一运行图是否在本运行图前侧链接
        :param reverse: 另一运行图是否转置
        :return:
        """
        if reverse:
            graph.reverse()

        if not line_only:
            # 车次连接
            for train_append in graph.trains():
                if self.checiExisted(train_append.fullCheci()):
                    # 本线有相同车次
                    train_main: Train = self.trainFromCheci(train_append.fullCheci())
                    train_main.delNonLocal(self)
                    train_append.delNonLocal(graph)
                    # 方向以本线为准
                    down = train_main.isDown(auto_guess=True, graph=self)
                    if down is None:
                        #如果本线无法判断，例如Z90终到石家庄在京石段只有一个站，则用另一条线的。
                        down = train_append.isDown(auto_guess=True,graph=graph)
                    if down is None:
                        #如果都无法判断，直接判断为下行车次
                        print("cannot judge down. use default.",train_main.fullCheci())
                        down = True

                    train_former = not (down ^ former)
                    train_main.jointTrain(train_append, train_former, graph)  # 当前站点已经拼接好

                else:
                    self.addTrain(train_append)

        #线路连接
        if former:
            for station in self.stationDicts():
                station["licheng"] += graph.lineLength()

            for st in reversed(graph.line.stations):
                if not self.stationExisted(st["zhanming"]):
                    self.line.addStation_by_origin(st,index=0)
        else:
            length = self.lineLength()
            for st in graph.stationDicts():
                st["licheng"] += length
                if not self.stationExisted(st["zhanming"]):
                    self.line.addStation_by_origin(st)



    def resetAllItems(self):
        for train in self.trains():
            train.setItem(None)

    def stationMile(self,name:str):
        """
        返回车站的里程数据，若不存在返回-1.不支持域解析符。2019.02.03删除线性算法。
        """
        st = self.line.stationDictByName(name,strict=True)
        if st is None:
            return -1
        return st["licheng"]

    def adjacentStation(self,name:str,ignore:list):
        index = self.stationIndex(name)
        if index>0:
            if self.line.stations[index-1] not in ignore:
                return self.line.stations[index-1]["zhanming"]

        if index<len(self.line.stations)-1:
            if self.line.stations[index+1] not in ignore:
                return self.line.stations[index+1]["zhanming"]

    def stationIndex(self,name:str):
        for i,st in enumerate(self.line.stations):
            if stationEqual(st["zhanming"] , name):
                return i
        raise Exception("No such station",name)

    def stationByDict(self,name:str,strict=False):
        """
        根据站名返回dict对象，函数名写错了。支持域解析符。
        2019.02.02删除线性算法。
        """
        return self.line.stationDictByName(name,strict)

    def resetStationName(self,old,new,auto_field=False):
        old_dict = self.stationByDict(old)
        if old_dict is not None:
            old_dict["zhanming"] = new if not auto_field else new.split('::')[0]

        for ruler in self.line.rulers:
            ruler.changeStationName(old,new)
        for train in self.trains():
            if train.isSfz(old):
                train.sfz=new
            if train.isZdz(old):
                train.zdz=new
            if train._localFirst == old:
                train._localFirst = new
            elif train._localLast == old:
                train._localLast = new
            st_dict = train.stationDict(old)
            if st_dict is not None:
                st_dict["zhanming"] = new
        self.line.changeStationNameUpdateMap(old,new)

    def addTrainByGraph(self,graph):
        """
        添加车次，返回数量
        """
        num = 0
        for train in graph.trains():
            if not self.checiExisted(train.fullCheci()):
                if train.localCount(self) >= 2:
                    num += 1
                    self.addTrain(train)
        return num

    def setMarkdown(self,mark:str):
        self._markdown = mark

    def markdown(self):
        try:
            return self._markdown
        except AttributeError:
            self._markdown = ""
            return ''

    def save_excel(self,filename:str):
        try:
            import openpyxl
            from openpyxl.styles import Font,Alignment
        except ImportError:
            return
        wb=openpyxl.Workbook()
        ws=wb.active
        ws['A1']=f'{self.firstStation()}-{self.lastStation()}间列车时刻表'

        #写入左侧表头
        ws['A3']='始发站'
        ws.merge_cells('A3:A4')
        ws['A5']='终到站'
        ws.merge_cells('A5:A6')
        ws['A7']='列车种类'
        ws.merge_cells('A7:A8')
        ws['A9']='车次'
        ws['A10']='车站'
        for row in range(3,9):
            ws.row_dimensions[row].font = Font(name='SimSum', size=9)
            ws.row_dimensions[row].alignment=Alignment(horizontal='center',vertical='center')
            ws.row_dimensions[row].height = 9.7

        start=11  #从第11行开始表格
        #写入车站
        station_row_dict={}
        cur=11
        for station in self.stations():
            ws.cell(row=cur,column=1,value=station)
            ws.merge_cells(start_row=cur,end_row=cur+1,start_column=1,end_column=1)
            station_row_dict[station]=cur
            ws.row_dimensions[cur].height=9.7
            ws.row_dimensions[cur+1].height=9.7
            ws.row_dimensions[cur].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[cur+1].alignment = Alignment(horizontal='center', vertical='center')
            cur += 2
        ws.column_dimensions['A'].width=12

        #写入车次，先下行
        last_merge_sfz,last_merge_zdz,last_merge_type=1,1,1
        col = 2
        last_train=None
        for train in self.trains():
            if not train.down:
                continue

            if last_train and train.sfz==last_train.sfz:
                try:
                    ws.unmerge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col - 1)
                except:
                    pass
                ws.merge_cells(start_row=3,end_row=4,start_column=last_merge_sfz,end_column=col)
            else:
                ws.merge_cells(start_row=3, end_row=4, start_column=col, end_column=col)
                last_merge_sfz=col
            ws.cell(row=3, column=col, value=train.sfz)

            if last_train and train.zdz == last_train.zdz:
                try:
                    ws.unmerge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col - 1)
                except:
                    pass
                ws.merge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col)
            else:
                ws.merge_cells(start_row=5, end_row=6, start_column=col, end_column=col)
                last_merge_zdz=col
            c=ws.cell(row=5, column=col, value=train.zdz)
            col_str=c.column
            ws.column_dimensions[col_str].width=6  #设置列宽为5

            if last_train and train.type == last_train.type:
                try:
                    ws.unmerge_cells(start_row=7,end_row=8,start_column=last_merge_type,end_column=col-1)
                except:
                    pass
                ws.merge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col)
            else:
                ws.merge_cells(start_row=7, end_row=8, start_column=col, end_column=col)
                last_merge_type=col
            ws.cell(row=7, column=col, value=train.type)

            checi=train.fullCheci()
            if '/' in checi:
                ws.cell(row=9,column=col,value=checi.split('/')[0])
                ws.cell(row=10,column=col,value='/'+checi.split('/',maxsplit=1)[1])
            else:
                ws.cell(row=9,column=col,value=checi)
                ws.merge_cells(start_row=9,end_row=10,start_column=col,end_column=col)

            last_dict=None

            # 时刻表循环
            for st_dict in train.timetable:
                for i,s in station_row_dict.items():
                    if stationEqual(i,st_dict['zhanming']):
                        row=s
                        break
                else:
                    continue

                if train.isSfz(st_dict['zhanming']):
                    ws.cell(row=row,column=col,value='')
                    ws.cell(row=row+1,column=col,value=self.outTime(st_dict['cfsj'],True))

                elif train.isZdz(st_dict["zhanming"]):
                    ws.cell(row=row,column=col,value=self.outTime(st_dict['ddsj'],True))
                    ws.cell(row=row+1,column=col,value='    --')

                elif train.stationStopped(st_dict):
                    #本站停车，无条件写入完整到达时刻和不完整出发时刻
                    ddsj_str=f'{st_dict["ddsj"].hour:2d}:{st_dict["ddsj"].minute:02d}'
                    sec=st_dict['ddsj'].second
                    if sec:
                        ddsj_str+=f"{sec:02d}"
                    else:
                        ddsj_str+='  '
                    ws.cell(row=row,column=col,value=ddsj_str)
                    if st_dict['ddsj'].hour==st_dict['cfsj'].hour:
                        cfsj_str='   '
                    else:
                        cfsj_str=f"{st_dict['cfsj'].hour:2d}:"
                    cfsj_str+=f'{st_dict["cfsj"].minute:02d}'
                    sec = st_dict['cfsj'].second
                    if sec:
                        ddsj_str += f"{sec:02d}"
                    else:
                        ddsj_str += '  '
                    ws.cell(row=row+1, column=col, value=cfsj_str)

                else:
                    give_hour=False
                    if not last_dict:
                        give_hour=True
                    elif last_dict['cfsj'].hour != st_dict['ddsj'].hour:
                        give_hour=True
                    ws.cell(row=row,column=col,value='   ...')
                    tgsj_str=f'{st_dict["ddsj"].hour:2d}:' if give_hour else '   '
                    tgsj_str+=f'{st_dict["ddsj"].minute:02d}'
                    sec = st_dict['ddsj'].second
                    if sec:
                        tgsj_str += f"{sec:02d}"
                    else:
                        tgsj_str += '  '
                    ws.cell(row=row+1, column=col, value=tgsj_str)
                last_dict=st_dict
            col+=1
            last_train=train

        #上行
        for train in self.trains():
            if train.down:
                continue
            if last_train and train.sfz==last_train.sfz:
                try:
                    ws.unmerge_cells(start_row=3, end_row=4, start_column=last_merge_sfz, end_column=col - 1)
                except:
                    pass
                ws.merge_cells(start_row=3,end_row=4,start_column=last_merge_sfz,end_column=col)
            else:
                ws.merge_cells(start_row=3, end_row=4, start_column=col, end_column=col)
                last_merge_sfz=col
            c=ws.cell(row=3, column=col, value=train.sfz)
            col_str = c.column
            ws.column_dimensions[col_str].width = 6  # 设置列宽为5

            if last_train and train.zdz == last_train.zdz:
                try:
                    ws.unmerge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col - 1)
                except:
                    pass
                ws.merge_cells(start_row=5, end_row=6, start_column=last_merge_zdz, end_column=col)
            else:
                ws.merge_cells(start_row=5, end_row=6, start_column=col, end_column=col)
                last_merge_zdz=col
            ws.cell(row=5, column=col, value=train.zdz)

            if last_train and train.type == last_train.type:
                try:
                    ws.unmerge_cells(start_row=7,end_row=8,start_column=last_merge_type,end_column=col-1)
                except:
                    pass
                ws.merge_cells(start_row=7, end_row=8, start_column=last_merge_type, end_column=col)
            else:
                ws.merge_cells(start_row=7, end_row=8, start_column=col, end_column=col)
                last_merge_type=col
            ws.cell(row=7, column=col, value=train.type)

            checi=train.fullCheci()
            if '/' in checi:
                ws.cell(row=9,column=col,value=checi.split('/')[0])
                ws.cell(row=10,column=col,value='/'+checi.split('/',maxsplit=1)[1])
            else:
                ws.cell(row=9,column=col,value=checi)
                ws.merge_cells(start_row=9,end_row=10,start_column=col,end_column=col)

            last_dict=None
            #时刻表循环
            for st_dict in train.timetable:
                for i,s in station_row_dict.items():
                    if stationEqual(i,st_dict['zhanming']):
                        row=s
                        break
                else:
                    continue

                if train.isSfz(st_dict['zhanming']):
                    ws.cell(row=row+1,column=col,value='')
                    ws.cell(row=row,column=col,value=self.outTime(st_dict['cfsj'],True))

                elif train.isZdz(st_dict["zhanming"]):
                    ws.cell(row=row+1,column=col,value=self.outTime(st_dict['ddsj'],True))
                    ws.cell(row=row,column=col,value='    --')

                elif train.stationStopped(st_dict):
                    #本站停车，无条件写入完整到达时刻和不完整出发时刻
                    ddsj_str=f'{st_dict["ddsj"].hour:2d}:{st_dict["ddsj"].minute:02d}'
                    sec=st_dict['ddsj'].second
                    if sec:
                        ddsj_str+=f"{sec:02d}"
                    else:
                        ddsj_str+='  '
                    ws.cell(row=row+1,column=col,value=ddsj_str)
                    if st_dict['ddsj'].hour==st_dict['cfsj'].hour:
                        cfsj_str='   '
                    else:
                        cfsj_str=f"{st_dict['cfsj'].hour:2d}:"
                    cfsj_str+=f'{st_dict["cfsj"].minute:02d}'
                    sec = st_dict['cfsj'].second
                    if sec:
                        ddsj_str += f"{sec:02d}"
                    else:
                        ddsj_str += '  '
                    ws.cell(row=row, column=col, value=cfsj_str)

                else:
                    give_hour=False
                    if not last_dict:
                        give_hour=True
                    elif last_dict['cfsj'].hour != st_dict['ddsj'].hour:
                        give_hour=True
                    ws.cell(row=row+1,column=col,value='   ...')
                    tgsj_str=f'{st_dict["ddsj"].hour:2d}:' if give_hour else '   '
                    tgsj_str+=f'{st_dict["ddsj"].minute:02d}'
                    sec = st_dict['ddsj'].second
                    if sec:
                        tgsj_str += f"{sec:02d}"
                    else:
                        tgsj_str += '  '
                    ws.cell(row=row, column=col, value=tgsj_str)
            col+=1
            last_train=train

        for row in range(1,ws.max_row+1):
            for col in range(1,ws.max_column+1):
                ws.cell(row=row,column=col).alignment=Alignment(horizontal='center',
                                                                vertical='center',shrink_to_fit=True)
                ws.cell(row=row,column=col).font=Font(name='宋体',size=9)


        wb.save(filename)

    def outTime(self,tgsj,give_hour:bool):
        tgsj_str = f'{tgsj.hour:2d}:' if give_hour else '   '
        tgsj_str += f'{tgsj.minute:02d}'
        sec = tgsj.second
        if sec:
            tgsj_str += f"{sec:02d}"
        else:
            tgsj_str += '  '
        return tgsj_str

    def getIntervalTrains(self,start,end,trainFilter):
        """
        返回某个区间办客车次列表。数据结构为list<dict>
        dict{
            'train':train object,
            'isSfz':boolean,
            'isZdz':boolean,
            'from':str,
            'to':str,
        """
        interval_list = []
        for train in self.trains():
            if not trainFilter.check(train):
                continue
            b1 = train.stationStopBehaviour(start)
            b2 = train.stationStopBehaviour(end)
            if '通过' in b1 or '通过' in b2:
                continue
            if not train.stationBefore(start,end):
                continue
            isSfz = (b1=='始发')
            isZdz = (b2=='终到')
            train_dict = {
                'train':train,
                'isSfz':isSfz,
                'isZdz':isZdz,
                'from':train.stationDict(start)['zhanming'],
                'to':train.stationDict(end)['zhanming']
            }
            interval_list.append(train_dict)
        return interval_list

    def getIntervalCount(self,fromOrTo, isStart, trainFilter):
        """
        获取区间对数表。
        :param fromOrTo:发站或到站
        :param isStart: True for start station, vice versa
        返回数据结构list<dict>
        dict{
        'from'
        """
        infoList = []
        if isStart:
            for st in self.stations():
                if not stationEqual(fromOrTo,st):
                    infoList.append({'from':fromOrTo,'to':st,'info':self.getIntervalTrains(fromOrTo,st,
                                                                                           trainFilter)})
        else:
            for st in self.stations():
                if not stationEqual(fromOrTo,st):
                    infoList.append({'to':fromOrTo,'from':st,'info':self.getIntervalTrains(st,fromOrTo,
                                                                                           trainFilter)})

        count_list = []
        for info_dict in infoList:
            info = info_dict['info']
            count = len(tuple(info))
            countSfz = len([1 for st in info if st['isSfz']])
            countZdz = len([1 for st in info if st['isZdz']])
            countSfZd = len([1 for st in info if st['isZdz'] and st['isSfz']])
            int_dict = {
                'from':info_dict['from'],
                'to':info_dict['to'],
                'count':count,
                'countSfz':countSfz,
                'countZdz':countZdz,
                'countSfZd':countSfZd
            }
            count_list.append(int_dict)
        return count_list

    def stationByIndex(self,idx):
        return self.line.stations[idx]

    def resetAllTrainsLocalFirstLast(self):
        """
        当线路数据更新时，重置所有列车的localFirst/Last。
        """
        for train in self.trains():
            train.updateLocalFirst(self)
            train.updateLocalLast(self)


if __name__ == '__main__':
    graph = Graph()
    graph.loadGraph("source/output.json")
    graph.show()
    graph.save("source/test.json")
